# -*- coding: utf-8 -*-
"""段2: ピッキング/梱包に「目標まで あと」「目標到達で 増やせる件数」列を追加する。

S列 = 目標まであと（文字。「1分28秒」／達成者は「達成済み」）
T列 = 目標到達で増やせる件数（数値。マイナスは目標ペースより先行している分）
Y列 = 短縮必要秒（隠し補助列。丸めを1か所に閉じ込め、分秒の桁ずれを防ぐ）
"""
import math, re, sys, zipfile
import openpyxl

SRC, DST, CALC = sys.argv[1], sys.argv[2], sys.argv[3]
SHEETS = {'xl/worksheets/sheet4.xml': ('ピッキング', 'C26'),
          'xl/worksheets/sheet5.xml': ('梱包', 'C27')}
HEAD = {'S': '目標まで\nあと', 'T': '目標到達で\n増やせる件数'}
STYLE = {'S': '50', 'T': '26'}          # 中央寄せ文字 / #,##0
NEWCOLS = [('S', 19), ('T', 20), ('Y', 25)]


def xesc(t):
    return t.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


def excel_round(x, n=0):
    """Excel の ROUND（0.5 は 0 から遠い側へ）。Python の round は銀行丸めなので使わない。"""
    f = 10 ** n
    return math.floor(abs(x) * f + 0.5) / f * (1 if x >= 0 else -1)


def fmt_gap(y):
    """短縮必要秒 -> 表示文字。Excel 側の数式と同じ結果になるようにする。"""
    if y <= 0:
        return '達成済み'
    y = int(y)
    if y < 60:
        return '%d秒' % y
    return '%d分%02d秒' % (y // 60, y % 60)


def num_cell(ref, style, formula, value, dec=None):
    if value is None:
        return ('<c r="%s"%s t="str"><f>%s</f><v/></c>'
                % (ref, style, formula))
    v = repr(value) if dec is None else ('%.*f' % (dec, value))
    return '<c r="%s"%s><f>%s</f><v>%s</v></c>' % (ref, style, formula, v)


def str_cell(ref, style, formula, text):
    body = '<v/>' if text is None else '<v>%s</v>' % xesc(text)
    return '<c r="%s"%s t="str"><f>%s</f>%s</c>' % (ref, style, formula, body)


def insert_cells(rowxml, newcells):
    """行XMLに新しいセルを列順で差し込む。"""
    m = re.match(r'(<row [^>]*>)(.*)(</row>)$', rowxml, re.S)
    head, body, tail = m.groups()
    cells = re.findall(r'<c r="[A-Z]+\d+"[^>]*/>|<c r="[A-Z]+\d+"[^>]*>.*?</c>', body, re.S)
    cells.extend(newcells)
    def key(c):
        col = re.match(r'<c r="([A-Z]+)', c).group(1)
        return openpyxl.utils.column_index_from_string(col)
    return head + ''.join(sorted(cells, key=key)) + tail


def main():
    calc = openpyxl.load_workbook(CALC, data_only=True)
    tgt_min = calc['設定']['C26'].value
    assert calc['設定']['C27'].value == tgt_min, '工程で目標が異なる場合は個別に読むこと'

    zin = zipfile.ZipFile(SRC)
    parts = {n: zin.read(n) for n in zin.namelist()}
    infos = {i.filename: i for i in zin.infolist()}
    order = zin.namelist()
    zin.close()

    for path, (name, tcell) in SHEETS.items():
        s = parts[path].decode('utf-8')
        ws = calc[name]
        tgt = calc['設定'][tcell].value
        tsec = tgt * 60

        # --- 列定義 ---
        assert '<col min="19"' not in s and '<col min="25"' not in s
        widths = {'S': '<col min="19" max="19" width="12" customWidth="1"/>',
                  'T': '<col min="20" max="20" width="14" customWidth="1"/>',
                  'Y': '<col min="25" max="25" width="9" hidden="1" customWidth="1"/>'}
        s = s.replace('<col min="22"', widths['S'] + widths['T'] + '<col min="22"', 1)
        s = s.replace('</cols>', widths['Y'] + '</cols>', 1)

        # --- 見出し ---
        heads = ['<c r="%s6" s="20" t="inlineStr"><is><t>%s</t></is></c>'
                 % (c, xesc(HEAD[c])) for c in ('S', 'T')]
        m = re.search(r'<row r="6"[^>]*>.*?</row>', s, re.S)
        s = s[:m.start()] + insert_cells(m.group(0), heads) + s[m.end():]

        # --- 明細行 ---
        f_y = 'IF(C{r}="","",ROUND(G{r}-設定!${c}*60,0))'
        f_s = ('IF(Y{r}="","",IF(Y{r}&lt;=0,"達成済み",IF(Y{r}&lt;60,Y{r}&amp;"秒",'
               'INT(Y{r}/60)&amp;"分"&amp;TEXT(MOD(Y{r},60),"00")&amp;"秒")))')
        f_t = 'IF(C{r}="","",D{r}/設定!${c}-C{r})'
        ck = tcell[0] + '$' + tcell[1:]      # C26 -> C$26

        n_live = 0
        for r in range(7, 67):
            C, D = ws.cell(r, 3).value, ws.cell(r, 4).value
            if C in (None, '') or not D:
                y = t = None
                text = None
            else:
                y = excel_round(D * 60 / C - tsec)
                text = fmt_gap(y)
                t = D / tgt - C
                n_live += 1
            cells = [
                str_cell('S%d' % r, ' s="%s"' % STYLE['S'], f_s.format(r=r), text),
                num_cell('T%d' % r, ' s="%s"' % STYLE['T'], f_t.format(r=r, c=ck), t),
                num_cell('Y%d' % r, '', f_y.format(r=r, c=ck), y),
            ]
            m = re.search(r'<row r="%d"[^>]*>.*?</row>' % r, s, re.S)
            s = s[:m.start()] + insert_cells(m.group(0), cells) + s[m.end():]

        parts[path] = s.encode('utf-8')
        print('%s: S/T/Y 列を60行ぶん追加（値の入る人 %d名・目標 %s分/件＝%.0f秒）'
              % (name, n_live, tgt, tsec))

    zout = zipfile.ZipFile(DST, 'w', zipfile.ZIP_DEFLATED)
    for nm in order:
        zi = zipfile.ZipInfo(nm, date_time=infos[nm].date_time)
        zi.compress_type = infos[nm].compress_type
        zi.external_attr = infos[nm].external_attr
        zout.writestr(zi, parts[nm])
    zout.close()


main()
