# -*- coding: utf-8 -*-
"""改良版サンプル: ③の帯を2段にして名前の重なりを解消 ＋ ②の全員版プレビュー。"""
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
import openpyxl

plt.rcParams["font.family"] = "IPAPGothic"
BLUE, GRAY, ORANGE, RED, GREEN, BROWN = "#2E75B6", "#7F7F7F", "#C55A11", "#C00000", "#1BAF7A", "#C55A11"

ws = openpyxl.load_workbook("cur.xlsx", data_only=True)["印刷用データ"]
both, only_pk = [], []
for r in range(3, 63):
    nm = ws[f"AG{r}"].value
    if not nm: continue
    pk, pack, proc = ws[f"AH{r}"].value, ws[f"AI{r}"].value, ws[f"AK{r}"].value
    (both if proc == "両方" else only_pk).append((nm, pk, pack))
only_pk = [(nm, pk) for nm, pk, _ in only_pk]
only_pk.sort(key=lambda t: t[1])

# ②用（良い順・全員）
bar = []
for r in range(3, 63):
    nm = ws[f"AP{r}"].value
    if not nm: continue
    bar.append((nm, ws[f"AQ{r}"].value, ws[f"AR{r}"].value))

fig = plt.figure(figsize=(19, 7.2))
gs = fig.add_gridspec(1, 2, width_ratios=[1.15, 1])

# ---------------- ③ 改良案A: 帯を2段に ----------------
ax = fig.add_subplot(gs[0, 0])
Y_LO, BAND_TOP = 10.0, 40.0
ax.set_xlim(40, 190); ax.set_ylim(Y_LO, 190)
ax.add_patch(Rectangle((40, Y_LO), 150, BAND_TOP - Y_LO, fc="#F2F2F2", ec="0.75", lw=0.9, zorder=0))
ax.text(188, BAND_TOP - 3, "梱包の実績なし（ピッキングのみ）", fontsize=7.5, color="0.3",
        ha="right", va="top", zorder=2)
ax.axvline(100, ymin=(BAND_TOP - Y_LO) / (190 - Y_LO), color="0.35", ls="--", lw=0.9)
ax.axhline(100, color="0.35", ls="--", lw=0.9)
for (fx, fy, t, c) in [(0.03, 0.96, "ピッキング改善", ORANGE), (0.72, 0.96, "良好", GREEN),
                       (0.03, 0.21, "要改善（両方）", RED), (0.72, 0.21, "梱包改善", ORANGE)]:
    ax.text(fx, fy, t, transform=ax.transAxes, fontsize=8, color=c, ha="left", va="center",
            bbox=dict(boxstyle="square,pad=0.35", fc="white", ec=c, lw=1), zorder=5)
for nm, x, y in both:
    ax.plot(x, y, "o", color=BLUE, ms=6, zorder=3)
    ax.annotate(nm, (x, y), textcoords="offset points", xytext=(0, 6), ha="center", fontsize=5.5)
for i, (nm, x) in enumerate(only_pk):                   # 2段に交互配置
    y = 30.0 if i % 2 == 0 else 19.0
    ax.plot(x, y, "^", color=GRAY, ms=6, zorder=3)
    ax.annotate(nm, (x, y), textcoords="offset points", xytext=(0, 5), ha="center", fontsize=5.5)
ax.set_xlabel("ピッキング効率（%）", fontsize=9); ax.set_ylabel("梱包効率（%）", fontsize=9)
ax.set_yticks([20, 40, 60, 80, 100, 120, 140, 160, 180])
ax.set_yticklabels(["", "40", "60", "80", "100", "120", "140", "160", "180"])
ax.grid(color="0.88", lw=0.6); ax.set_axisbelow(True); ax.tick_params(labelsize=8)
ax.set_title("③ 改良案A  帯を2段にして名前の重なりを解消（灰色▲＝ピッキングのみ10名）", fontsize=11)

# ---------------- ② 全員版 ----------------
ax2 = fig.add_subplot(gs[0, 1])
names = [b[0] for b in bar]
xs = range(len(bar))
w = 0.4
ax2.bar([i - w/2 for i in xs], [b[1] or 0 for b in bar], w, color=BLUE, label="ピッキング効率")
ax2.bar([i + w/2 for i in xs], [b[2] or 0 for b in bar], w, color=BROWN, label="梱包効率")
ax2.axhline(100, color=RED, ls="--", lw=1, label="目標100%")
ax2.set_xticks(list(xs)); ax2.set_xticklabels(names, rotation=60, ha="right", fontsize=6)
ax2.set_ylabel("効率（%）", fontsize=9); ax2.legend(fontsize=7, ncol=3, loc="upper right")
ax2.grid(axis="y", color="0.88", lw=0.6); ax2.set_axisbelow(True); ax2.tick_params(labelsize=8)
ax2.set_title("② 全員29名版（梱包の棒が無い人＝ピッキングのみ）", fontsize=11)

fig.tight_layout()
fig.savefig("sample2.png", dpi=110)
print("wrote sample2.png")
print("PKのみ:", [n for n, _ in only_pk])
