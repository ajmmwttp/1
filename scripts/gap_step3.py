# -*- coding: utf-8 -*-
"""段3+段4: 新しい2列を表（テーブル）と判定の行色に含め、集計ブロックに全体行を足す。"""
import re, sys, zipfile
import openpyxl

SRC, DST, CALC = sys.argv[1], sys.argv[2], sys.argv[3]
SHEETS = {'xl/worksheets/sheet4.xml': ('ピッキング', 'C$26', 'xl/tables/table1.xml'),
          'xl/worksheets/sheet5.xml': ('梱包', 'C$27', 'xl/tables/table2.xml')}
NEW_COLS = [(19, '目標まで&#10;あと'), (20, '目標到達で&#10;増やせる件数')]
# 集計ブロックに足す行: (行, 見出し, 数式, スタイル)
SUMMARY = [
    (18, '1件あたり 目標まで(秒)', 'IFERROR(W17-設定!${c}*60,"")', '22'),
    (19, '目標到達で増やせる件数', 'IFERROR(W12/設定!${c}-W10,"")', '26'),
]


def insert_cells(rowxml, newcells):
    m = re.match(r'(<row [^>]*>)(.*)(</row>)$', rowxml, re.S)
    head, body, tail = m.groups()
    cells = re.findall(r'<c r="[A-Z]+\d+"[^>]*/>|<c r="[A-Z]+\d+"[^>]*>.*?</c>', body, re.S)
    cells.extend(newcells)
    return head + ''.join(sorted(cells, key=lambda c: openpyxl.utils.column_index_from_string(
        re.match(r'<c r="([A-Z]+)', c).group(1)))) + tail


def main():
    calc = openpyxl.load_workbook(CALC, data_only=True)
    zin = zipfile.ZipFile(SRC)
    parts = {n: zin.read(n) for n in zin.namelist()}
    infos = {i.filename: i for i in zin.infolist()}
    order = zin.namelist()
    zin.close()

    for path, (name, ck, tpath) in SHEETS.items():
        s = parts[path].decode('utf-8')
        ws = calc[name]

        # --- 判定の行色を新列まで広げる ---
        assert '<conditionalFormatting sqref="A7:R66">' in s, '%s: CF範囲が想定と違う' % name
        s = s.replace('<conditionalFormatting sqref="A7:R66">',
                      '<conditionalFormatting sqref="A7:T66">', 1)

        # --- 集計ブロックの全体行 ---
        added = []
        for row, label, f, style in SUMMARY:
            fml = f.replace('${c}', '$' + ck)
            val = (ws['W17'].value - 96 if row == 18
                   else ws['W12'].value / 1.6 - ws['W10'].value)
            cells = [
                '<c r="V%d" s="47" t="inlineStr"><is><t>%s</t></is></c>' % (row, label),
                '<c r="W%d" s="%s"><f>%s</f><v>%s</v></c>' % (row, style, fml, repr(val)),
            ]
            m = re.search(r'<row r="%d"[^>]*>.*?</row>' % row, s, re.S)
            s = s[:m.start()] + insert_cells(m.group(0), cells) + s[m.end():]
            added.append('%s = %.1f' % (label, val))
        parts[path] = s.encode('utf-8')

        # --- テーブル範囲と列定義 ---
        t = parts[tpath].decode('utf-8')
        assert t.count('A6:R66') == 2, '%s: テーブル範囲が想定と違う' % tpath
        t = t.replace('A6:R66', 'A6:T66')
        t = t.replace('<tableColumns count="18">', '<tableColumns count="20">')
        extra = ''.join('<tableColumn id="%d" name="%s"/>' % (i, n) for i, n in NEW_COLS)
        t = t.replace('</tableColumns>', extra + '</tableColumns>')
        parts[tpath] = t.encode('utf-8')

        print('%s: 判定の行色を A7:T66 へ / テーブルを A6:T66・20列へ / 集計行 %s'
              % (name, ' ・ '.join(added)))

    zout = zipfile.ZipFile(DST, 'w', zipfile.ZIP_DEFLATED)
    for nm in order:
        zi = zipfile.ZipInfo(nm, date_time=infos[nm].date_time)
        zi.compress_type = infos[nm].compress_type
        zi.external_attr = infos[nm].external_attr
        zout.writestr(zi, parts[nm])
    zout.close()


main()
