# -*- coding: utf-8 -*-
"""段1+3: 全シートの数式セルのキャッシュを真値に書き換え、fullCalcOnLoad を立てる。

前回の納品で、書き換えたセルの下流のキャッシュが古いまま残り、Excelが
全再計算しない設定だったため誤った値が表示されていた。数式には触れず、
保存されている計算結果（<v>）だけを正しい値に入れ替える。
"""
import re, sys, zipfile
import openpyxl
from openpyxl.utils import get_column_letter as CL

SRC, DST, TRUTH = sys.argv[1], sys.argv[2], sys.argv[3]


def xesc(t):
    return (t.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;'))


def render(v):
    """真値 -> (t属性, <v>の中身)。None は空文字結果として扱う。"""
    if v is None:
        return ' t="str"', None
    if isinstance(v, bool):
        return ' t="b"', '1' if v else '0'
    if isinstance(v, (int, float)):
        return '', repr(v) if isinstance(v, float) else str(v)
    s = str(v)
    if s.startswith('#') and s.endswith(('!', '?', 'A', '0')):
        return ' t="e"', xesc(s)
    if s == '':
        return ' t="str"', None
    return ' t="str"', xesc(s)


def main():
    truth = openpyxl.load_workbook(TRUTH, data_only=True)
    zin = zipfile.ZipFile(SRC)
    parts = {n: zin.read(n) for n in zin.namelist()}
    infos = {i.filename: i for i in zin.infolist()}
    order = zin.namelist()
    zin.close()

    # workbook.xml のシート名 -> worksheet パス
    wbx = parts['xl/workbook.xml'].decode('utf-8')
    rels = parts['xl/_rels/workbook.xml.rels'].decode('utf-8')
    rid2tgt = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels))
    sheetmap = {}
    for m in re.finditer(r'<sheet name="([^"]+)"[^>]*r:id="(rId\d+)"', wbx):
        tgt = rid2tgt[m.group(2)]
        sheetmap[m.group(1)] = 'xl/' + tgt.lstrip('/').replace('worksheets/', 'worksheets/')

    total = 0
    for name, path in sheetmap.items():
        if path not in parts:
            path = 'xl/worksheets/' + path.split('/')[-1]
        s = parts[path].decode('utf-8')
        ws = truth[name]
        n = 0
        out = []
        pos = 0
        # 空セルの自己終了タグ <c .../> を巻き込まないよう (?<!/) で除外する
        for m in re.finditer(r'<c r="([A-Z]+)(\d+)"([^>]*)(?<!/)>((?:(?!</c>).)*)</c>', s, re.S):
            body = m.group(4)
            if '<f' not in body:
                continue
            col, row, attrs = m.group(1), int(m.group(2)), m.group(3)
            val = ws.cell(row, openpyxl.utils.column_index_from_string(col)).value
            t, txt = render(val)
            attrs = re.sub(r' t="[^"]*"', '', attrs) + t
            mf = re.search(r'<f[^>]*/>|<f[^>]*>(?:(?!</f>).)*</f>', body, re.S)
            newbody = mf.group(0) + ('<v/>' if txt is None else '<v>%s</v>' % txt)
            out.append(s[pos:m.start()])
            out.append('<c r="%s%d"%s>%s</c>' % (col, row, attrs, newbody))
            pos = m.end()
            n += 1
        out.append(s[pos:])
        parts[path] = ''.join(out).encode('utf-8')
        total += n
        if n:
            print('  %s: %d セルのキャッシュを更新' % (name, n))

    print('合計 %d セル' % total)

    # 再発防止: 開くたびに全再計算させる
    m = re.search(r'<calcPr([^>]*)/>', wbx)
    assert m, 'calcPr がない'
    if 'fullCalcOnLoad' in m.group(1):
        print('workbook.xml: fullCalcOnLoad は設定済み')
    else:
        parts['xl/workbook.xml'] = (wbx[:m.end()-2] + ' fullCalcOnLoad="1"/>' + wbx[m.end():]).encode('utf-8')
        print('workbook.xml: fullCalcOnLoad="1" を設定')

    zout = zipfile.ZipFile(DST, 'w', zipfile.ZIP_DEFLATED)
    for nm in order:
        zi = zipfile.ZipInfo(nm, date_time=infos[nm].date_time)
        zi.compress_type = infos[nm].compress_type
        zi.external_attr = infos[nm].external_attr
        zout.writestr(zi, parts[nm])
    zout.close()


main()
