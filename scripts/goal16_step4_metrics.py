# -*- coding: utf-8 -*-
"""段4: 会社目標(1.6分/件)との比較値を 印刷用データ に追加する。

追加する集計（すべて数式なので入力を替えれば自動追随）:
  BL2  会社目標(分/件)          = 設定!C26
  BL3  PK 総件数 / BL4 総企業数 / BL5 総実測(分)
  BL6  PK 会社目標達成率(%)      = 目標分×件数 ÷ 実測 ×100
  BL7  PK 1社あたり件数
  BL8  PK 段取りが占める秒/件    = a × 企業数/件数
  BL9  PK 1.6分に必要な1社件数   = a ÷ (96 − b)
  BL11 梱包 総件数 / BL12 総実測 / BL13 梱包 会社目標達成率
"""
import re
import zipfile

import openpyxl

SRC, DST = "s3.xlsx", "s4.xlsx"
D, LASTROW = "印刷用データ", 62

# ---- キャッシュ値を Python で計算（新係数 a=56.05 で）----
wv = openpyxl.load_workbook(SRC, data_only=True)
st = wv["設定"]
A, B = 56.05, st["E7"].value          # 変更後の a / 据え置きの b
GOAL_MIN = st["C26"].value            # 1.6
GOAL_SEC = GOAL_MIN * 60              # 96.0

def agg(sheet, col):
    ws = wv[sheet]
    return sum(ws.cell(r, col).value for r in range(7, 67)
               if isinstance(ws.cell(r, col).value, (int, float)))

pk_cnt, pk_comp, pk_act = agg("ピッキング", 3), agg("ピッキング", 5), agg("ピッキング", 4)
pa_cnt, pa_act = agg("梱包", 3), agg("梱包", 4)

V = {
    2:  GOAL_MIN,
    3:  pk_cnt,
    4:  pk_comp,
    5:  pk_act,
    6:  GOAL_MIN * pk_cnt / pk_act * 100,
    7:  pk_cnt / pk_comp,
    8:  A * pk_comp / pk_cnt,
    9:  A / (GOAL_SEC - B),
    11: pa_cnt,
    12: pa_act,
    13: GOAL_MIN * pa_cnt / pa_act * 100,
}
LABEL = {
    2: "会社目標(分/件)", 3: "PK 総件数", 4: "PK 総企業数", 5: "PK 総実測(分)",
    6: "PK 会社目標達成率(%)", 7: "PK 1社あたり件数", 8: "PK 段取り(秒/件)",
    9: "PK 1.6分に必要な1社件数", 11: "梱包 総件数", 12: "梱包 総実測(分)",
    13: "梱包 会社目標達成率(%)",
}
FORMULA = {
    2:  "設定!$C$26",
    3:  "SUM(ピッキング!$C$7:$C$66)",
    4:  "SUM(ピッキング!$E$7:$E$66)",
    5:  "SUM(ピッキング!$D$7:$D$66)",
    6:  'IF($BL$5=0,"",$BL$2*$BL$3/$BL$5*100)',
    7:  'IF($BL$4=0,"",$BL$3/$BL$4)',
    8:  'IF($BL$3=0,"",設定!$E$6*$BL$4/$BL$3)',
    9:  'IF(設定!$E$26-設定!$E$7&lt;=0,"",設定!$E$6/(設定!$E$26-設定!$E$7))',
    11: "SUM(梱包!$C$7:$C$66)",
    12: "SUM(梱包!$D$7:$D$66)",
    13: 'IF($BL$12=0,"",設定!$C$27*$BL$11/$BL$12*100)',
}
print("=== Python で算出した値（新係数 a=56.05）===")
for k in sorted(V):
    print(f"  BL{k:<3} {LABEL[k]:<26} = {V[k]:,.4f}")
print()
print(f"  ピッキング: 会社目標 {GOAL_MIN}分×{pk_cnt:,.0f}件 = {GOAL_MIN*pk_cnt:,.0f}分 / 実測 {pk_act:,.0f}分")
print(f"  梱包      : 会社目標 {GOAL_MIN}分×{pa_cnt:,.0f}件 = {GOAL_MIN*pa_cnt:,.0f}分 / 実測 {pa_act:,.0f}分")

def patch(x):
    assert '<c r="BK' not in x, "BK 列が既にある"
    for r in sorted(set(LABEL) | set(V)):
        cells = (f'<c r="BK{r}" s="39" t="inlineStr"><is><t>{LABEL[r]}</t></is></c>'
                 f'<c r="BL{r}" s="62"><f>{FORMULA[r]}</f><v>{V[r]!r}</v></c>')
        m = re.search(r'(<row r="%d"[ >].*?)(</row>)' % r, x, re.S)
        assert m, f"行{r} が無い"
        x = x[:m.start(2)] + cells + x[m.start(2):]
    x = x.replace('<dimension ref="A1:BI62"/>', '<dimension ref="A1:BL62"/>', 1)
    x = x.replace('spans="1:61"', 'spans="1:64"')
    x = x.replace('<col min="1" max="61" width="11" customWidth="1"/>',
                  '<col min="1" max="64" width="11" customWidth="1"/>', 1)
    return x

zi = zipfile.ZipFile(SRC); zo = zipfile.ZipFile(DST, "w", zipfile.ZIP_DEFLATED)
for it in zi.infolist():
    d = zi.read(it.filename)
    if it.filename == "xl/worksheets/sheet12.xml":
        d = patch(d.decode("utf-8")).encode("utf-8")
    zo.writestr(it, d)
zo.close(); zi.close()
print("\nwrote", DST)
