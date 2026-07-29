import openpyxl, sys
a = openpyxl.load_workbook('work.xlsx', read_only=True, data_only=True)
b = openpyxl.load_workbook('集計表_集計済み.xlsx', read_only=True, data_only=True)
assert a.sheetnames == b.sheetnames, (a.sheetnames, b.sheetnames)
print('sheetnames identical:', len(a.sheetnames), 'sheets')
DATES=[s for s in a.sheetnames if s.isdigit()]
diff=0; cells=0
for name in DATES + ['Sheet1']:
    ra=a[name].iter_rows(values_only=True); rb=b[name].iter_rows(values_only=True)
    for i,(x,y) in enumerate(zip(ra,rb), start=1):
        if x!=y:
            diff+=1
            if diff<=5: print('DIFF', name, 'row', i)
        cells+=len(x)
print('date-sheet rows compared, differing rows:', diff, '| cells:', cells)
# 集計表 A列 1-35 unchanged
wa=a['集計表']; wbk=b['集計表']
oa=[c[0] for c in wa.iter_rows(min_row=1,max_row=35,max_col=1,values_only=True)]
ob=[c[0] for c in wbk.iter_rows(min_row=1,max_row=35,max_col=1,values_only=True)]
print('集計表 A1:A35 unchanged:', oa==ob)
a.close(); b.close()
# formula lengths
wf=openpyxl.load_workbook('集計表_集計済み.xlsx')['集計表']
L=[len(c.value) for r in wf.iter_rows() for c in r if isinstance(c.value,str) and c.value.startswith('=')]
print('formula cells:', len(L), 'max len:', max(L), '(Excel limit 8192)')
assert max(L) < 8192
print('sample B2:', wf['B2'].value[:150], '...')
sys.exit(1 if diff else 0)
