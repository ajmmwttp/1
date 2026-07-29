"""生成した集計表の数式を Python で解釈・評価し、生データから独立に求めた正解と突き合わせる。

LibreOffice に依存せず「実際に書き込まれた数式文字列」を検証する。
COUNTIF/COUNTIFS は Excel の照合規則（ワイルドカード * ? 、"=" は空セル）を再現する。
"""
import re, sys, collections
import openpyxl

SRC = 'work.xlsx'
OUT = sys.argv[1] if len(sys.argv) > 1 else 'out_wild.xlsx'
MAXROW = 1000
CATS = ['バラ検品', 'バラ検品即', '梱包']

# ---------- 生データ読み込み（C列=作業区分, E列=作業者名） ----------
wbd = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
DATES = [s for s in wbd.sheetnames if s.isdigit()]
sheet_data = {}
for s in wbd.worksheets:
    if s.title not in DATES:
        continue
    C, E = [], []
    for i, r in enumerate(s.iter_rows(values_only=True), start=1):
        if i == 1 or i > MAXROW:
            continue
        C.append(r[2]); E.append(r[4])
    # 2..MAXROW 行に満たない分は空セルで埋める
    pad = (MAXROW - 1) - len(C)
    C += [None] * pad; E += [None] * pad
    sheet_data[s.title] = (C, E)
wbd.close()

# ---------- 集計表の数式と、A列の文字列 ----------
wbf = openpyxl.load_workbook(OUT)
wsf = wbf['集計表']
literal = {}   # 座標 -> 文字列/数値（数式でないセル）
for row in wsf.iter_rows(min_row=1, max_row=wsf.max_row, max_col=5):
    for c in row:
        if c.value is not None and not (isinstance(c.value, str) and c.value.startswith('=')):
            literal[c.coordinate] = c.value


# ---------- Excel の COUNTIF 照合規則 ----------
def excel_match(cell, crit):
    if crit == '=':                      # 空セル
        return cell is None or cell == ''
    if cell is None:
        return False
    s = str(cell)
    if any(ch in crit for ch in '*?'):
        rx = '^' + ''.join(
            '.*' if ch == '*' else '.' if ch == '?' else re.escape(ch)
            for ch in crit) + '$'
        return re.match(rx, s, re.DOTALL) is not None
    return s.casefold() == crit.casefold()


# ---------- 数式評価 ----------
REF = re.compile(r'^\$?([A-Z]+)\$?(\d+)$')
SUBST = re.compile(r'^SUBSTITUTE\((\$?[A-Z]+\$?\d+),"(.*?)","(.*?)"\)$')


def crit_value(expr):
    """COUNTIF(S) の検索条件式を実際の文字列に評価する。"""
    expr = expr.strip()
    if expr.startswith('"') and expr.endswith('"'):
        return expr[1:-1]
    m = SUBST.match(expr)
    if m:
        base = crit_value(m.group(1))
        return base.replace(m.group(2), m.group(3))
    m = REF.match(expr)
    if m:
        v = literal.get(f'{m.group(1)}{m.group(2)}')
        if v is None:
            raise ValueError(f'参照先が空: {expr}')
        return str(v)
    raise ValueError(f'未対応の条件式: {expr}')


def split_args(s):
    """括弧・引用符を考慮してカンマで分割する。"""
    out, depth, q, cur = [], 0, False, ''
    for ch in s:
        if ch == '"':
            q = not q
        if not q:
            if ch == '(':
                depth += 1
            elif ch == ')':
                depth -= 1
            elif ch == ',' and depth == 0:
                out.append(cur); cur = ''; continue
        cur += ch
    out.append(cur)
    return out


def split_top(s, ops='+-'):
    """トップレベルの + / - で分割し [(符号, 項), ...] を返す。"""
    out, depth, q, cur, sign = [], 0, False, '', 1
    for ch in s:
        if ch == '"':
            q = not q
        if not q:
            if ch == '(':
                depth += 1
            elif ch == ')':
                depth -= 1
            elif depth == 0 and ch in ops and cur:
                out.append((sign, cur)); cur = ''; sign = 1 if ch == '+' else -1
                continue
        cur += ch
    out.append((sign, cur))
    return out


RANGE = re.compile(r"^'([^']+)'!\$([CE])\$2:\$([CE])\$(\d+)$")


def eval_term(t):
    t = t.strip()
    if t.startswith('COUNTIFS(') and t.endswith(')'):
        a = split_args(t[len('COUNTIFS('):-1])
        assert len(a) == 4, a
        m1, m2 = RANGE.match(a[0].strip()), RANGE.match(a[2].strip())
        assert m1 and m2 and m1.group(1) == m2.group(1), a
        assert int(m1.group(4)) == MAXROW and int(m2.group(4)) == MAXROW, a
        C, E = sheet_data[m1.group(1)]
        col1 = C if m1.group(2) == 'C' else E
        col2 = C if m2.group(2) == 'C' else E
        c1, c2 = crit_value(a[1]), crit_value(a[3])
        return sum(1 for x, y in zip(col1, col2)
                   if excel_match(x, c1) and excel_match(y, c2))
    if t.startswith('COUNTIF(') and t.endswith(')'):
        a = split_args(t[len('COUNTIF('):-1])
        assert len(a) == 2, a
        m = RANGE.match(a[0].strip())
        C, E = sheet_data[m.group(1)]
        col = C if m.group(2) == 'C' else E
        c = crit_value(a[1])
        return sum(1 for x in col if excel_match(x, c))
    if t.startswith('SUM(') and t.endswith(')'):
        inner = t[4:-1]
        a, b = inner.split(':')
        m1, m2 = REF.match(a), REF.match(b)
        c1 = openpyxl.utils.column_index_from_string(m1.group(1))
        c2 = openpyxl.utils.column_index_from_string(m2.group(1))
        r1, r2 = int(m1.group(2)), int(m2.group(2))
        tot = 0
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                tot += evaluate(f'{openpyxl.utils.get_column_letter(cc)}{rr}')
        return tot
    m = REF.match(t)
    if m:
        return evaluate(f'{m.group(1)}{m.group(2)}')
    raise ValueError(f'未対応の項: {t[:80]}')


_cache = {}


def evaluate(coord):
    if coord in _cache:
        return _cache[coord]
    v = wsf[coord].value
    if v is None:
        return 0
    if not (isinstance(v, str) and v.startswith('=')):
        return v
    total = sum(sign * eval_term(term) for sign, term in split_top(v[1:]))
    _cache[coord] = total
    return total


# ---------- 独立に求めた正解 ----------
recs = []
for d in DATES:
    C, E = sheet_data[d]
    recs += list(zip(C, E))
nrm = lambda s: re.sub(r'[\s　]+', '', s or '')
truth = collections.Counter()
for c, n in recs:
    if c is not None:
        truth[(nrm(n), c)] += 1

# ---------- 突き合わせ ----------
LABELS = {'名簿計', '名簿外計', '合計（名簿＋名簿外）', '作業者名が空欄の実績',
          '実績データ全件（検算用）', '差異（0ならOK）'}
fails, checked = [], 0
name_rows = [r for r in range(2, wsf.max_row + 1)
             if isinstance(wsf.cell(r, 1).value, str)
             and wsf.cell(r, 1).value not in LABELS
             and isinstance(wsf.cell(r, 2).value, str)
             and wsf.cell(r, 2).value.startswith('=COUNTIFS')]
print(f'--- 氏名行 {len(name_rows)} 行を検証 ---')
print(f"{'':22s}{'バラ検品':>9s}{'バラ検品即':>11s}{'梱包':>7s}{'合計':>7s}")
for r in name_rows:
    nm = wsf.cell(r, 1).value
    got = [evaluate(f'{c}{r}') for c in 'BCD']
    exp = [truth[(nrm(nm), cat)] for cat in CATS]
    checked += 3
    if got != exp:
        fails.append((r, nm, got, exp))
    tot = evaluate(f'E{r}')
    assert tot == sum(got), (r, tot, got)
    print(f'{nm:22s}{got[0]:9d}{got[1]:11d}{got[2]:7d}{tot:7d}')

print('\n--- 集計・検算行 ---')
for r in range(2, wsf.max_row + 1):
    lab = wsf.cell(r, 1).value
    if isinstance(lab, str) and lab in ('名簿計', '名簿外計', '合計（名簿＋名簿外）',
                                        '作業者名が空欄の実績', '実績データ全件（検算用）',
                                        '差異（0ならOK）'):
        vals = [evaluate(f'{c}{r}') for c in 'BCDE']
        print(f'{lab:22s}' + ''.join(f'{v:>9d}' for v in vals))
        if lab == '差異（0ならOK）' and any(vals):
            fails.append((r, lab, vals, [0, 0, 0, 0]))
    if isinstance(lab, str) and lab in ('バラPK即', 'バラPK'):
        print(f'{lab:22s}{evaluate(f"B{r}"):>9d}')

raw = collections.Counter(c for c, _ in recs if c is not None)
print('\n生データの区分別件数:', {c: raw[c] for c in CATS + ['バラPK即', 'バラPK']})
print(f'\n検証セル数: {checked}  不一致: {len(fails)}')
for f in fails:
    print('  NG', f)
sys.exit(1 if fails else 0)
