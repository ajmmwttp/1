# -*- coding: utf-8 -*-
"""段2: ピッキング/梱包に U列「目標到達で減らせる人件費(円)」を追加する。

U = (実測時間(分) − 件数×目標1.6分) ÷ 60 × 時給（設定!$C$28）。
未達者はプラス（目標到達で減らせる額）、達成者はマイナス（既に目標より安く処理している額）。
"""
import re, sys, zipfile
import openpyxl

SRC, DST, CALC = sys.argv[1], sys.argv[2], sys.argv[3]
SHEETS = {'xl/worksheets/sheet4.xml': ('ピッキング', 'C$26'),
          'xl/worksheets/sheet5.xml': ('梱包', 'C$27')}
HEADER = '目標到達で\n減らせる人件費(円)'
WAGE = 1500


def xesc(t):
    return t.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


def insert_cells(rowxml, newcells):
    m = re.match(r'(<row [^>]*>)(.*)(</row>)$', rowxml, re.S)
    head, body, tail = m.groups()
    cells = re.findall(r'<c r="[A-Z]+\d+"[^>]*/>|<c r="[A-Z]+\d+"[^>]*>.*?</c>', body, re.S)
    cells.extend(newcells)
    return head + ''.join(sorted(cells, key=lambda c: openpyxl.utils.column_index_from_string(
        re.match(r'<c r="([A-Z]+)', c).group(1)))) + tail


def main():
    calc = openpyxl.load_workbook(CALC, data_only=True)

    zin = zipfile.ZipFile(SRC)
    parts = {n: zin.read(n) for n in zin.namelist()}
    infos = {i.filename: i for i in zin.infolist()}
    order = zin.namelist()
    zin.close()

    for path, (name, ck) in SHEETS.items():
        s = parts[path].decode('utf-8')
        ws = calc[name]
        tgt = calc['設定'][ck.replace('$', '')].value

        # 列幅（21列目は未定義のはず）
        assert '<col min="21"' not in s, '%s: 21列目の定義が既にある' % name
        s = s.replace('<col min="22"',
                      '<col min="21" max="21" width="16" customWidth="1"/><col min="22"', 1)

        # 見出し
        assert '<c r="U6"' not in s, '%s: U6 が既にある' % name
        head = '<c r="U6" s="20" t="inlineStr"><is><t>%s</t></is></c>' % xesc(HEADER)
        m = re.search(r'<row r="6"[^>]*>.*?</row>', s, re.S)
        s = s[:m.start()] + insert_cells(m.group(0), [head]) + s[m.end():]

        # 明細
        fml = 'IF(C{r}="","",(D{r}-C{r}*設定!${c})/60*設定!$C$28)'
        n_live = 0
        for r in range(7, 67):
            C, D = ws.cell(r, 3).value, ws.cell(r, 4).value
            if C in (None, '') or D in (None, ''):
                cell = ('<c r="U%d" s="26" t="str"><f>%s</f><v/></c>'
                        % (r, fml.format(r=r, c=ck)))
            else:
                v = (D - C * tgt) / 60 * WAGE
                cell = ('<c r="U%d" s="26"><f>%s</f><v>%s</v></c>'
                        % (r, fml.format(r=r, c=ck), repr(v)))
                n_live += 1
            m = re.search(r'<row r="%d"[^>]*>.*?</row>' % r, s, re.S)
            s = s[:m.start()] + insert_cells(m.group(0), [cell]) + s[m.end():]

        parts[path] = s.encode('utf-8')
        print('%s: U列を60行ぶん追加（値の入る人 %d名）' % (name, n_live))

    zout = zipfile.ZipFile(DST, 'w', zipfile.ZIP_DEFLATED)
    for nm in order:
        zi = zipfile.ZipInfo(nm, date_time=infos[nm].date_time)
        zi.compress_type = infos[nm].compress_type
        zi.external_attr = infos[nm].external_attr
        zout.writestr(zi, parts[nm])
    zout.close()


main()
