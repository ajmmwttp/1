# -*- coding: utf-8 -*-
"""②内訳比較グラフの改善案を、実際の描画領域と同じ縦横比で比較する実験。

②の領域は 列T:AA × 行5-16 = 約 5.08in × 3.5in。
50名分（最悪ケース）のデータで各案を描く。
"""
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl

plt.rcParams["font.family"] = "IPAPGothic"
BLUE, BROWN, RED, GRAY = "#2E75B6", "#C55A11", "#C00000", "#808080"
W, H = 5.08, 3.5                      # ②の実寸(inch)

ws = openpyxl.load_workbook("chk_50.xlsx", data_only=True)["印刷用データ"]
good = []                              # 良い順（AP/AQ/AR）
for r in range(3, 63):
    nm = ws[f"AP{r}"].value
    if nm:
        good.append((nm, ws[f"AQ{r}"].value, ws[f"AR{r}"].value))
worst = list(reversed(good))           # 改善優先＝悪い順
print(f"対象 {len(good)} 名")

def frame(ax, title):
    ax.set_title(title, fontsize=8.5, pad=4)
    ax.axhline(100, color=RED, ls="--", lw=0.9, zorder=1)
    ax.grid(axis="y", color="0.88", lw=0.5); ax.set_axisbelow(True)
    ax.tick_params(labelsize=5)
    ax.set_ylabel("効率（%）", fontsize=6)

def grouped(ax, data, labels=False, title=""):
    xs = range(len(data)); w = 0.4
    ax.bar([i-w/2 for i in xs], [d[1] or 0 for d in data], w, color=BLUE, label="ピッキング")
    ax.bar([i+w/2 for i in xs], [d[2] or 0 for d in data], w, color=BROWN, label="梱包")
    ax.set_xticks(list(xs)); ax.set_xticklabels([d[0] for d in data], rotation=60, ha="right",
                                                fontsize=5 if len(data) > 25 else 6)
    if labels:
        for i, d in enumerate(data):
            for off, v in ((-w/2, d[1]), (w/2, d[2])):
                if v: ax.text(i+off, v+3, f"{v:.0f}", ha="center", fontsize=4, rotation=90)
    frame(ax, title); ax.legend(fontsize=5, ncol=3, loc="upper right")

# ---- 4案 ----
fig, axes = plt.subplots(2, 2, figsize=(W*2.15, H*2.15))

grouped(axes[0][0], good, False, "案① 現状のまま 50名（1本 3.7pt）")
grouped(axes[0][1], worst[:20], True, "案② 改善優先の上位20名だけ（値ラベル付き）")

# 案③ ダンベル（全員・1人1行）
ax = axes[1][0]
d3 = worst
ys = range(len(d3))
for i, (nm, pk, pk2) in enumerate(d3):
    a, b = pk or 0, pk2 or 0
    if pk and pk2: ax.plot([a, b], [i, i], color="0.75", lw=0.8, zorder=1)
    if pk:  ax.plot(a, i, "o", color=BLUE,  ms=2.5, zorder=2)
    if pk2: ax.plot(b, i, "o", color=BROWN, ms=2.5, zorder=2)
ax.axvline(100, color=RED, ls="--", lw=0.9)
ax.set_yticks(list(ys)); ax.set_yticklabels([d[0] for d in d3], fontsize=4)
ax.set_ylim(-1, len(d3)); ax.invert_yaxis()
ax.set_xlabel("効率（%）", fontsize=6); ax.grid(axis="x", color="0.88", lw=0.5)
ax.set_axisbelow(True); ax.tick_params(labelsize=5)
ax.set_title("案③ ダンベル 全員50名（●ピッキング ●梱包）", fontsize=8.5, pad=4)

# 案④ 差分棒（梱包効率 − ピッキング効率）
ax = axes[1][1]
d4 = [(nm, (b or 0) - (a or 0), a, b) for nm, a, b in worst]
xs = range(len(d4))
cols = [BROWN if v[1] < 0 else BLUE for v in d4]
ax.bar(list(xs), [v[1] for v in d4], 0.7,
       color=[GRAY if (v[2] is None or v[3] is None) else c for v, c in zip(d4, cols)])
ax.axhline(0, color="0.3", lw=0.8)
ax.set_xticks(list(xs)); ax.set_xticklabels([v[0] for v in d4], rotation=60, ha="right", fontsize=5)
ax.set_ylabel("梱包 − ピッキング（pt）", fontsize=6)
ax.grid(axis="y", color="0.88", lw=0.5); ax.set_axisbelow(True); ax.tick_params(labelsize=5)
ax.set_title("案④ 差分1本 全員50名（青=梱包が速い / 茶=ピッキングが速い）", fontsize=8.5, pad=4)

fig.suptitle("② ピッキング・梱包 内訳比較 — 50名での見え方 4案（実領域と同じ縦横比）", fontsize=11)
fig.tight_layout(rect=[0, 0, 1, 0.95])
fig.savefig("exp_chart2.png", dpi=115)
print("wrote exp_chart2.png")
