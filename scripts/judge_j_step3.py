# -*- coding: utf-8 -*-
"""段3: O列の判定を4区分（会社目標達成率・点推定）に差し替える。

旧: 縮小後効率の95%CIが群平均T3をまたぐか（◎上位/○標準/△要支援/日数未入力）
新: R列（目標達成率%）としきい値（設定!D31/D32/D33）による4区分
    ◎達成(≥100) / ○あと一歩(≥90) / △要改善(≥80) / ✕要支援(<80)
共有数式グループ（si=5: O7:O38, si=14: O39:O66）を個別式に展開する。
"""
import re
import zipfile

import openpyxl

SRC, DST = "j2.xlsx", "j3.xlsx"

# ---- キャッシュ: R列の値（j1で算出済み）から判定を Python で計算 ----
wv = openpyxl.load_workbook(SRC, data_only=True)
st = wv["設定"]
HI, MID, LO = st["D31"].value, st["D32"].value, st["D33"].value
assert (HI, MID, LO) == (100, 90, 80)

def judge(v):
    if v is None: return None
    return ("◎達成" if v >= HI else "○あと一歩" if v >= MID
            else "△要改善" if v >= LO else "✕要支援")

CACHE = {}
from collections import Counter
for sn, nm in [("sheet4", "ピッキング"), ("sheet5", "梱包")]:
    ws = wv[nm]
    CACHE[sn] = {r: judge(ws.cell(r, 18).value) for r in range(7, 67)
                 if isinstance(ws.cell(r, 18).value, (int, float))}
    print(f"{nm}: {dict(Counter(CACHE[sn].values()))}")

def patch(x, sn):
    # 共有数式のマスター2つと従属セルをすべて個別式へ
    n_repl = 0
    for r in range(7, 67):
        f = (f'IF(R{r}="","",IF(R{r}&gt;=設定!$D$31,"◎達成",'
             f'IF(R{r}&gt;=設定!$D$32,"○あと一歩",'
             f'IF(R{r}&gt;=設定!$D$33,"△要改善","✕要支援"))))')
        v = CACHE[sn].get(r)
        new = (f'<c r="O{r}" s="50" t="str"><f>{f}</f><v>{v}</v></c>' if v
               else f'<c r="O{r}" s="50" t="str"><f>{f}</f><v/></c>')
        m = re.search(r'<c r="O%d"[^>]*>.*?</c>' % r, x, re.S)
        assert m, f"O{r} が無い"
        x = x.replace(m.group(0), new, 1)
        n_repl += 1
    assert n_repl == 60
    assert 'si="5"' not in re.search(r'<row r="7"[ >].*?</row>', x, re.S).group(0).split('<c r="O7"')[1].split('</c>')[0]
    return x

zi = zipfile.ZipFile(SRC); zo = zipfile.ZipFile(DST, "w", zipfile.ZIP_DEFLATED)
for it in zi.infolist():
    b = zi.read(it.filename)
    if it.filename in ("xl/worksheets/sheet4.xml", "xl/worksheets/sheet5.xml"):
        b = patch(b.decode("utf-8"), it.filename.split("/")[-1][:-4]).encode("utf-8")
    zo.writestr(it, b)
zo.close(); zi.close()
print("wrote", DST)
