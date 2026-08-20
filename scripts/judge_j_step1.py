# -*- coding: utf-8 -*-
"""段1: ピッキング/梱包に R列「目標達成率%」を追加し、fullCalcOnLoad を立てる。

fullCalcOnLoad: アップロード版は Excel が再保存したのに再計算しておらず
（calcId が古いまま一致するため）、係数変更前のキャッシュが画面と印刷に
出ていた。次回オープン時に全再計算させるフラグを workbook に立てる。
"""
import re
import zipfile

import openpyxl

SRC, DST = "upload27.xlsx", "j1.xlsx"

# ---- キャッシュ値: アップロード版の入力データから Python で算出 ----
wv = openpyxl.load_workbook(SRC, data_only=True)
inp = wv["入力"]
GOAL = {"sheet4": ("C26", 3, 4), "sheet5": ("C27", 7, 8)}   # 目標セル, 件数列, 時間列
goal_min = {k: wv["設定"][v[0]].value for k, v in GOAL.items()}
CACHE = {}
for sn, (gref, ccol, tcol) in GOAL.items():
    g = wv["設定"][gref].value
    vals = {}
    for i in range(60):                       # 入力 行10..69 → シート 行7..66
        r_in, r_sh = 10 + i, 7 + i
        c, t = inp.cell(r_in, ccol).value, inp.cell(r_in, tcol).value
        if all(isinstance(v, (int, float)) and v > 0 for v in (c, t)):
            vals[r_sh] = g * c / t * 100
    CACHE[sn] = vals
print("キャッシュ算出: PK", len(CACHE["sheet4"]), "名 / 梱包", len(CACHE["sheet5"]), "名")

def patch_sheet(x, sn):
    gcell = "$C$26" if sn == "sheet4" else "$C$27"
    # 1) 見出し R6（O6 と同じ s=20）
    assert '<c r="R6"' not in x
    m6 = re.search(r'(<c r="Q6"[^>]*>.*?</c>)', x)
    assert m6, "Q6 が無い"
    x = x.replace(m6.group(1), m6.group(1)
                  + '<c r="R6" s="20" t="inlineStr"><is><t>目標達成率\n%</t></is></c>', 1)
    # 2) データ行 7..66
    for i in range(60):
        r = 7 + i
        f = f'IF(C{r}="","",設定!{gcell}*C{r}/D{r}*100)'
        v = CACHE[sn].get(r)
        cell = (f'<c r="R{r}" s="22"><f>{f}</f><v>{v!r}</v></c>' if v is not None
                else f'<c r="R{r}" s="22" t="str"><f>{f}</f><v/></c>')
        mq = re.search(r'(<c r="Q%d"[^>]*(?:/>|>.*?</c>))' % r, x)
        assert mq, f"Q{r} が無い"
        # R セルは Q の直後（S より前）に挿入
        x = x.replace(mq.group(1), mq.group(1) + cell, 1)
    # 3) 縞模様の条件付き書式を R 列まで拡張
    x = x.replace('sqref="A7:Q66"', 'sqref="A7:R66"', 1)
    # 4) R 列の幅（K 列などと同じ 9）
    m = re.search(r'(<col min="15" max="(\d+)"[^>]*/>)', x)
    cols = re.search(r'<cols>(.*?)</cols>', x, re.S).group(1)
    if 'min="18"' not in cols:
        # 17列目までの定義の後に列18を追加（既存定義とぶつからない形で）
        tail = re.search(r'<col min="(\d+)" max="(\d+)"([^>]*)/></cols>', x)
        x = x.replace('</cols>', '<col min="18" max="18" width="10" customWidth="1"/></cols>', 1)
    return x

def patch_workbook(x):
    old = '<calcPr calcId="191029"/>'
    assert old in x, "calcPr が想定と違う"
    return x.replace(old, '<calcPr calcId="191029" fullCalcOnLoad="1"/>', 1)

PATCH = {"xl/worksheets/sheet4.xml": lambda s: patch_sheet(s, "sheet4"),
         "xl/worksheets/sheet5.xml": lambda s: patch_sheet(s, "sheet5"),
         "xl/workbook.xml": patch_workbook}
zi = zipfile.ZipFile(SRC); zo = zipfile.ZipFile(DST, "w", zipfile.ZIP_DEFLATED)
for it in zi.infolist():
    b = zi.read(it.filename)
    fn = PATCH.get(it.filename)
    if fn:
        b = fn(b.decode("utf-8")).encode("utf-8")
    zo.writestr(it, b)
zo.close(); zi.close()
print("wrote", DST)
