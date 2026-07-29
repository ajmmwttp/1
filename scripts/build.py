"""集計表に バラ検品 / バラ検品即 / 梱包 の件数を Excel 関数で集計する。

MODE:
  'wildcard' … 氏名の全角スペースを * に置換した 1 条件で COUNTIFS（21項/セル）
  'exact'    … 全角スペース版と半角スペース3個版の 2 条件を加算（42項/セル・ワイルドカード非依存）
"""
import re, sys, fnmatch, collections
from copy import copy
import openpyxl
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter

SRC = 'work.xlsx'
OUT = sys.argv[1] if len(sys.argv) > 1 else 'out.xlsx'
MODE = sys.argv[2] if len(sys.argv) > 2 else 'wildcard'
MAXROW = 1000
CATS = ['バラ検品', 'バラ検品即', '梱包']
OTHER = ['バラPK即', 'バラPK']
IDSP = '　'      # 全角スペース
ASCII3 = '   '       # 半角スペース3個（実績データ側の「河野   忠」用）

wb = openpyxl.load_workbook(SRC)
ws = wb['集計表']
DATES = [s for s in wb.sheetnames if s.isdigit()]
assert len(DATES) == 21, DATES

# ---------- 1. 実績データ側の氏名を収集 ----------
wbv = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
roster = [ws.cell(row=r, column=1).value for r in range(2, 36)]
norm = lambda s: re.sub(r'[\s　]+', '', s or '')

data_names = collections.Counter()
for s in wbv.worksheets:
    if s.title not in DATES:
        continue
    rows = s.iter_rows(values_only=True)
    next(rows)
    for r in rows:
        if r[4] is not None:
            data_names[r[4]] += 1
wbv.close()

rnorm = {norm(n) for n in roster}
offroster = sorted({n for n in data_names if norm(n) not in rnorm},
                   key=lambda n: (-data_names[n], n))
allnames = sorted(data_names)

# ---------- 2. 照合方式の安全性を検証 ----------
assert not any(ch in n for n in roster + offroster for ch in '*?~'), 'ワイルドカード文字を含む氏名'
for n in roster + offroster:
    # wildcard: 誤マッチが起きないこと
    hits = [m for m in allnames if fnmatch.fnmatchcase(m, n.replace(IDSP, '*'))]
    assert len(hits) <= 1, f'ワイルドカード衝突: {n!r} -> {hits}'
    # exact: 全角版と半角3個版が両方データに存在する（＝二重計上する）氏名が無いこと
    assert not (n in data_names and n.replace(IDSP, ASCII3) in data_names), f'二重計上: {n!r}'
# 実績データの氏名が「全角スペース区切り」か「半角スペース3個区切り」のどちらかに収まること
covered = {n for n in roster + offroster} | {n.replace(IDSP, ASCII3) for n in roster + offroster}
assert set(allnames) <= covered, f'未対応の表記ゆれ: {set(allnames) - covered}'

# ---------- 3. レイアウト ----------
R_ROSTER = (2, 35)
R_ROSTER_TOT = 36
R_OFF_HDR = 38
R_OFF = (39, 39 + len(offroster) - 1)
R_OFF_TOT = R_OFF[1] + 1
R_TOTAL = R_OFF_TOT + 2
R_NONAME = R_TOTAL + 1
R_ALL = R_TOTAL + 2
R_DIFF = R_TOTAL + 3
R_REF_HDR = R_DIFF + 2
R_REF = R_REF_HDR + 1
R_NOTE = R_REF + len(OTHER) + 1
COLS = ['B', 'C', 'D']


def name_refs(r):
    """氏名の照合条件（1個または2個）を返す。"""
    if MODE == 'wildcard':
        return [f'SUBSTITUTE($A{r},"{IDSP}","*")']
    return [f'$A{r}', f'SUBSTITUTE($A{r},"{IDSP}","{ASCII3}")']


def sum_countifs(cat_ref, refs):
    """21枚の日付シートを横断して COUNTIFS を合計する。"""
    return '=' + '+'.join(
        f"COUNTIFS('{d}'!$C$2:$C${MAXROW},{cat_ref},'{d}'!$E$2:$E${MAXROW},{nr})"
        for d in DATES for nr in refs)


def sum_countif(cat_ref):
    return '=' + '+'.join(f"COUNTIF('{d}'!$C$2:$C${MAXROW},{cat_ref})" for d in DATES)


# ---------- 4. スタイル（既存書式に合わせる） ----------
hdr_font = copy(ws['A1'].font)      # 游ゴシック 11
body_font = copy(ws['A2'].font)     # ＭＳ Ｐゴシック 11
bold_font = copy(body_font); bold_font.b = True
note_font = copy(body_font); note_font.sz = 9
thin = Side(style='thin')
box = Border(left=thin, right=thin, top=thin, bottom=thin)
NUM = '#,##0'


def put(row, col, value, *, font, border=True, numfmt=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = copy(font)
    if border:
        c.border = copy(box)
    if numfmt:
        c.number_format = numfmt
    if align:
        c.alignment = Alignment(horizontal=align, vertical='center')
    return c


def data_row(r, refs):
    for i, col in enumerate(COLS):
        put(r, 2 + i, sum_countifs(f'{col}$1', refs), font=body_font, numfmt=NUM)
    put(r, 5, f'=SUM(B{r}:D{r})', font=body_font, numfmt=NUM)


def total_row(r, label, formula):
    put(r, 1, label, font=bold_font, align='left')
    for c in range(2, 6):
        put(r, c, formula(get_column_letter(c)), font=bold_font, numfmt=NUM)


# ---------- 5. ヘッダー ----------
for i, name in enumerate(CATS + ['合計']):
    put(1, 2 + i, name, font=hdr_font, align='center')

# ---------- 6. 名簿の34名 ----------
for r in range(R_ROSTER[0], R_ROSTER[1] + 1):
    data_row(r, name_refs(r))
total_row(R_ROSTER_TOT, '名簿計', lambda L: f'=SUM({L}{R_ROSTER[0]}:{L}{R_ROSTER[1]})')

# ---------- 7. 名簿外の作業者 ----------
put(R_OFF_HDR, 1, '【名簿外】実績はあるが上記名簿に未記載の作業者',
    font=bold_font, border=False, align='left')
for j, nm in enumerate(offroster):
    r = R_OFF[0] + j
    put(r, 1, nm, font=body_font, align='left')
    data_row(r, name_refs(r))
total_row(R_OFF_TOT, '名簿外計', lambda L: f'=SUM({L}{R_OFF[0]}:{L}{R_OFF[1]})')

# ---------- 8. 合計と検算 ----------
total_row(R_TOTAL, '合計（名簿＋名簿外）', lambda L: f'={L}{R_ROSTER_TOT}+{L}{R_OFF_TOT}')

put(R_NONAME, 1, '作業者名が空欄の実績', font=body_font, align='left')
for i, col in enumerate(COLS):
    put(R_NONAME, 2 + i, sum_countifs(f'{col}$1', ['"="']), font=body_font, numfmt=NUM)
put(R_NONAME, 5, f'=SUM(B{R_NONAME}:D{R_NONAME})', font=body_font, numfmt=NUM)

put(R_ALL, 1, '実績データ全件（検算用）', font=body_font, align='left')
for i, col in enumerate(COLS):
    put(R_ALL, 2 + i, sum_countif(f'{col}$1'), font=body_font, numfmt=NUM)
put(R_ALL, 5, f'=SUM(B{R_ALL}:D{R_ALL})', font=body_font, numfmt=NUM)

put(R_DIFF, 1, '差異（0ならOK）', font=body_font, align='left')
for c in range(2, 6):
    L = get_column_letter(c)
    put(R_DIFF, c, f'={L}{R_ALL}-{L}{R_TOTAL}-{L}{R_NONAME}', font=body_font, numfmt=NUM)

# ---------- 9. 参考：集計対象外の作業区分 ----------
put(R_REF_HDR, 1, '【参考】今回の集計対象外の作業区分', font=bold_font, border=False, align='left')
for j, cat in enumerate(OTHER):
    r = R_REF + j
    put(r, 1, cat, font=body_font, align='left')
    put(r, 2, sum_countif(f'$A{r}'), font=body_font, numfmt=NUM)

# ---------- 10. 注記 ----------
match_note = (
    f'※氏名の照合は COUNTIFS。「河野{IDSP}忠」のみ実績データ側が半角スペース区切り（"河野{ASCII3}忠"）のため、'
    + ('氏名の全角スペースを * に置換して照合しています。'
       if MODE == 'wildcard' else
       '全角スペース版と半角スペース3個版の2条件を加算しています。')
)
notes = [
    '※集計対象：本ブックの日付シート21枚（20260401〜20260430）。各シートの C列「作業区分」と E列「作業者名」を参照。',
    '※数値は実績データの行数（レコード件数）です。数量（PK数量・総数）の合計ではありません。',
    match_note,
    '　（全氏名について誤マッチ・二重計上が起きないことを確認済みです。）',
    '※参照範囲は各シートの2〜1000行目。行が増えても拾えるよう余裕を持たせています（現在の最大は539行）。',
]
for j, t in enumerate(notes):
    put(R_NOTE + j, 1, t, font=note_font, border=False, align='left')

# ---------- 11. 列幅 ----------
ws.column_dimensions['A'].width = 24   # 名簿外の長い氏名に合わせて拡幅
for L in 'BCDE':
    ws.column_dimensions[L].width = 11

# ---------- 12. 開いたときに必ず再計算させる ----------
# openpyxl は数式のキャッシュ値を書けないため、読み込み時の全再計算を指定しておく。
wb.calculation.fullCalcOnLoad = True

wb.save(OUT)
print(f'saved {OUT} (MODE={MODE})')
print(f'offroster {len(offroster)}: {offroster}')
print(f'rows: roster_tot={R_ROSTER_TOT} off={R_OFF} off_tot={R_OFF_TOT} '
      f'total={R_TOTAL} noname={R_NONAME} all={R_ALL} diff={R_DIFF} ref={R_REF} note={R_NOTE}')
