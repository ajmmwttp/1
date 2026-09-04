# -*- coding: utf-8 -*-
"""段6: グラフの参照元を Excel でも正しく描ける形にする。

監査（グラフ観点）で分かったこと:
  a) ダッシュボード①②（chart1/2）の氏名ラベルが「点の番号ごとの固定文字」で、
     入力シート順のまま。点はピッキング表（並べ替え済み）の行順なので、
     29 名中 28 名のラベルが別人に付いていた。
     → ③（chart10）と同じ「1人1系列・系列名＝氏名セル」方式に組み替える
       （①用 AH 列・②用 AM 列に氏名を数式で置き、系列名で参照。標準の OOXML なので
       Excel でも LibreOffice でも描ける）。ラベル位置は衝突を避けて振り直す。
  b) 散布図 8 枚（chart1/2/11〜16）の X 値が <c:strRef>（文字列扱い）で保存されて
     いた。Excel は X を文字とみなすと 1,2,3… の位置に描く。→ <c:numRef> に戻す。
  c) 参照元の空行が "" を返すため、Excel では線グラフの目標100%線が 0 に落ち
     （印刷用① chart8・② chart9）、散布図では (0,0) に点が出る。
     → グラフ専用の値は NA() を返す（#N/A は Excel も LibreOffice も描かない）。
       ダッシュボード AF/AG/AK/AL・印刷用データ AV/BI はグラフ専用なので式を直接
       NA() に。散布図の AE〜BL は相関などの統計にも使うので触らず、グラフ専用の
       複製列 BR〜CD（=IF(ISNUMBER(x),x,NA())）を足して chart11〜16 をそちらに向ける。

キャッシュは後段（LibreOffice 再計算 → cache_step1/2）で更新する。
"""
import re, sys, zipfile
import openpyxl

SRC, DST, VALS = sys.argv[1], sys.argv[2], sys.argv[3]
C15 = 'http://schemas.microsoft.com/office/drawing/2012/chart'


def esc(t):
    return t.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


def cell_re(ref):
    return re.compile(r'<c r="%s"(?:(?: [^>]*?)?/>|(?: [^>]*)?>(?:(?!</c>).)*</c>)' % ref, re.S)


def na_in_hosts(s, cols, rows, old, new, label):
    """列 cols の共有数式ホスト（式の本文を持つセル）の old → new。"""
    n = 0
    for col in cols:
        for r in rows:
            m = cell_re('%s%d' % (col, r)).search(s)
            if not m:
                continue
            blk = m.group(0)
            mf = re.search(r'(<f[^>]*>)((?:(?!</f>).)*)(</f>)', blk, re.S)
            if not mf or old not in mf.group(2):
                continue
            blk2 = blk[:mf.start(2)] + mf.group(2).replace(old, new) + blk[mf.end(2):]
            s = s[:m.start()] + blk2 + s[m.end():]
            n += 1
    print('%s: 式 %d 件を NA() に' % (label, n))
    assert n >= len(cols), label
    return s


def insert_after(s, after_ref, new_cell_xml):
    m = cell_re(after_ref).search(s)
    assert m, after_ref + ' が無い'
    return s[:m.end()] + new_cell_xml + s[m.end():]


def main():
    wv = openpyxl.load_workbook(VALS, data_only=True)
    zin = zipfile.ZipFile(SRC)
    parts = {n: zin.read(n) for n in zin.namelist()}
    infos = {i.filename: i for i in zin.infolist()}
    order = zin.namelist()
    zin.close()

    # ------------------------------------------------ ダッシュボード
    s = parts['xl/worksheets/sheet1.xml'].decode('utf-8')
    s = na_in_hosts(s, ('AF', 'AG', 'AK', 'AL'), range(3, 63), ',"")', ',NA())', 'ダッシュボード AF/AG/AK/AL')
    # 氏名列 AH（①）/ AM（②）
    s = insert_after(s, 'AG2', '<c r="AH2" s="3" t="inlineStr"><is><t>① 氏名</t></is></c>')
    s = insert_after(s, 'AL2', '<c r="AM2" s="3" t="inlineStr"><is><t>② 氏名</t></is></c>')
    for r in range(3, 63):
        s = insert_after(s, 'AG%d' % r,
            '<c r="AH%d" t="str"><f>IFERROR(INDEX(ピッキング!$B$7:$B$66,MATCH(T%d,$AE$3:$AE$62,0)),"")</f><v/></c>' % (r, r))
        s = insert_after(s, 'AL%d' % r,
            '<c r="AM%d" t="str"><f>IFERROR(INDEX(ピッキング!$B$7:$B$66,MATCH(T%d,$AJ$3:$AJ$62,0)),"")</f><v/></c>' % (r, r))
    parts['xl/worksheets/sheet1.xml'] = s.encode('utf-8')
    print('ダッシュボード: AH/AM に①②の氏名列（行3〜62）を追加')

    # ラベル用の氏名（キャッシュ）: 対番号 k の人 = ピッキング表の k 番目に両値が揃う行
    ws = wv['ダッシュボード']; pk = wv['ピッキング']
    def names_for(seq_col):
        out = []
        for k in range(1, 61):
            nm = ''
            for i in range(60):
                if ws.cell(3 + i, seq_col).value == k:
                    nm = pk.cell(7 + i, 2).value or ''
                    break
            out.append(nm)
        return out
    names1, names2 = names_for(31), names_for(36)          # AE=31, AJ=36
    print('   ①の氏名: %s … / ②の氏名: %s …' % (names1[:3], names2[:3]))

    # ------------------------------------------------ 印刷用データ AV / BI
    s = parts['xl/worksheets/sheet12.xml'].decode('utf-8')
    s = na_in_hosts(s, ('AV',), (3, 35), '"",100)', 'NA(),100)', '印刷用データ AV（① 目標線）')
    s = na_in_hosts(s, ('BI',), (3,), '"",100)', 'NA(),100)', '印刷用データ BI（② 目標線）')
    parts['xl/worksheets/sheet12.xml'] = s.encode('utf-8')

    # ------------------------------------------------ 散布図 グラフ専用列 BR〜CD
    s = parts['xl/worksheets/sheet7.xml'].decode('utf-8')
    SRC_COLS = ['AE', 'AF', 'AI', 'AJ', 'AM', 'AN', 'AQ', 'AR', 'BE', 'BF', 'BJ', 'BK', 'BL']
    NEW_COLS = ['BR', 'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 'BY', 'BZ', 'CA', 'CB', 'CC', 'CD']
    LABELS = ['①x', '①y', '②x', '②y', '③x', '③y', '④x', '④y', '⑤x', '⑤y', '⑥x', '⑥y', '⑥z']
    def append_cells(s, r, cells):
        m = re.search(r'(<row r="%d"[^>]*>)(.*?)(</row>)' % r, s, re.S)
        assert m, 'row %d が無い' % r
        head = re.sub(r'spans="(\d+):\d+"', r'spans="\1:82"', m.group(1))
        return s[:m.start()] + head + m.group(2) + cells + m.group(3) + s[m.end():]
    hdr = ''.join('<c r="%s7" s="3" t="inlineStr"><is><t>グラフ用 %s</t></is></c>' % (c, l)
                  for c, l in zip(NEW_COLS, LABELS))
    s = append_cells(s, 7, hdr)
    for r in range(8, 68):
        cells = ''.join('<c r="%s%d"><f>IF(ISNUMBER(%s%d),%s%d,NA())</f><v>0</v></c>' % (nc, r, sc, r, sc, r)
                        for nc, sc in zip(NEW_COLS, SRC_COLS))
        s = append_cells(s, r, cells)
    s = s.replace('<dimension ref="B2:BP155"/>', '<dimension ref="B2:CD155"/>', 1)
    s = s.replace('</cols>', '<col min="70" max="82" width="13" customWidth="1"/></cols>', 1)
    parts['xl/worksheets/sheet7.xml'] = s.encode('utf-8')
    print('散布図: グラフ専用列 BR〜CD（行8〜67、空は NA()）を追加')

    # ------------------------------------------------ チャート
    def to_numref(c, remap):
        n = 0
        def repl(m):
            nonlocal n
            ref = m.group(1)
            for a, b in remap.items():
                ref = ref.replace(a, b)
            n += 1
            return ('<c:xVal><c:numRef><c:f>%s</c:f><c:numCache><c:formatCode>General</c:formatCode>'
                    '<c:ptCount val="60"/></c:numCache></c:numRef></c:xVal>' % ref)
        c = re.sub(r'<c:xVal><c:strRef><c:f>([^<]*)</c:f>(?:(?!</c:xVal>).)*</c:xVal>', repl, c, flags=re.S)
        # yVal / 残りの参照も付け替え（数値参照はそのまま）
        for a, b in remap.items():
            c = c.replace('<c:f>%s</c:f>' % a, '<c:f>%s</c:f>' % b)
        return c, n

    def sheet_ref(sh, col):
        return '%s!$%s$8:$%s$67' % (sh, col, col)
    REMAP = {11: ('AE', 'BR', 'AF', 'BS'), 12: ('AI', 'BT', 'AJ', 'BU'), 13: ('AM', 'BV', 'AN', 'BW'),
             14: ('AQ', 'BX', 'AR', 'BY'), 15: ('BE', 'BZ', 'BF', 'CA')}
    for n, (x0, x1, y0, y1) in REMAP.items():
        p = 'xl/charts/chart%d.xml' % n
        c = parts[p].decode('utf-8')
        c, k = to_numref(c, {sheet_ref('散布図', x0): sheet_ref('散布図', x1),
                             sheet_ref('散布図', y0): sheet_ref('散布図', y1)})
        assert k == 1 and sheet_ref('散布図', x1) in c and sheet_ref('散布図', y1) in c, p
        parts[p] = c.encode('utf-8')
    c = parts['xl/charts/chart16.xml'].decode('utf-8')
    c, k = to_numref(c, {sheet_ref('散布図', 'BJ'): sheet_ref('散布図', 'CB'),
                         sheet_ref('散布図', 'BK'): sheet_ref('散布図', 'CC'),
                         sheet_ref('散布図', 'BL'): sheet_ref('散布図', 'CD')})
    assert k == 2, 'chart16'
    parts['xl/charts/chart16.xml'] = c.encode('utf-8')
    print('chart11〜16: X 値を numRef に戻し、参照をグラフ専用列へ')

    # ---- chart1/2 を 1人1系列に組み替える（③ chart10 と同じ方式） ----
    XR, YR = (0.0, 0.8), (0.0, 200.0)          # 軸範囲（固定）
    PW, PH = 420.0, 280.0                     # プロット領域の見積り(pt)
    def lbl_box(x, y, pos, nm):
        cx, cy = (x - XR[0]) / (XR[1] - XR[0]) * PW, (y - YR[0]) / (YR[1] - YR[0]) * PH
        w, h, g, mr = len(nm) * 8.0, 9.6, 2.5, 3.5
        if pos == 't': return (cx - w / 2, cy + mr + g, cx + w / 2, cy + mr + g + h)
        if pos == 'b': return (cx - w / 2, cy - mr - g - h, cx + w / 2, cy - mr - g)
        if pos == 'r': return (cx + mr + g, cy - h / 2, cx + mr + g + w, cy + h / 2)
        return (cx - mr - g - w, cy - h / 2, cx - mr - g, cy + h / 2)
    def mk_box(x, y):
        cx, cy = (x - XR[0]) / (XR[1] - XR[0]) * PW, (y - YR[0]) / (YR[1] - YR[0]) * PH
        return (cx - 3.5, cy - 3.5, cx + 3.5, cy + 3.5)
    def hit(a, b):
        return not (a[2] <= b[0] or b[2] <= a[0] or a[3] <= b[1] or b[3] <= a[1])
    def solve(pts):
        markers = {r: mk_box(x, y) for r, (nm, x, y) in pts.items()}
        placed, out = {}, {}
        for r in sorted(pts, key=lambda k: (-pts[k][2], pts[k][1])):
            nm, x, y = pts[r]
            for pos in ('r', 'l', 't', 'b'):
                b = lbl_box(x, y, pos, nm)
                if b[0] < -1 or b[2] > PW + 1 or b[1] < -1 or b[3] > PH + 1: continue
                if any(hit(b, mb) for k, mb in markers.items() if k != r): continue
                if any(hit(b, pb) for pb in placed.values()): continue
                out[r] = pos; break
            else:
                out[r] = 'r'
            placed[r] = lbl_box(x, y, out[r], nm)
        return out

    def rebuild(c, name_col, x_col, y_col, seq_col, names):
        ser0 = re.search(r'<c:ser>.*?</c:ser>', c, re.S)
        marker = re.search(r'<c:marker>.*?</c:marker>', ser0.group(0), re.S).group(0)
        pts = {}
        for r in range(3, 63):
            from openpyxl.utils import column_index_from_string as ci
            x, y = ws.cell(r, ci(x_col)).value, ws.cell(r, ci(y_col)).value
            if isinstance(x, (int, float)) and isinstance(y, (int, float)) and names[r - 3]:
                pts[r] = (names[r - 3], float(x), float(y))
        pos = solve(pts)
        body = []
        for i, r in enumerate(range(3, 63)):
            nm = names[r - 3]
            cache = ('<c:strCache><c:ptCount val="1"/><c:pt idx="0"><c:v>%s</c:v></c:pt></c:strCache>' % esc(nm)) if nm else ''
            body.append(
                '<c:ser><c:idx val="%d"/><c:order val="%d"/>' % (i, i)
                + '<c:tx><c:strRef><c:f>ダッシュボード!$%s$%d</c:f>%s</c:strRef></c:tx>' % (name_col, r, cache)
                + '<c:spPr><a:ln><a:noFill/><a:prstDash val="solid"/></a:ln></c:spPr>' + marker
                + '<c:dLbls><c:spPr><a:noFill/><a:ln><a:noFill/></a:ln></c:spPr>'
                  '<c:txPr><a:bodyPr/><a:lstStyle/><a:p><a:pPr><a:defRPr sz="800"/></a:pPr><a:endParaRPr lang="ja-JP"/></a:p></c:txPr>'
                  '<c:dLblPos val="%s"/><c:showLegendKey val="0"/><c:showVal val="0"/><c:showCatName val="0"/>'
                  '<c:showSerName val="1"/><c:showPercent val="0"/><c:showBubbleSize val="0"/><c:showLeaderLines val="0"/></c:dLbls>' % pos.get(r, 'r')
                + '<c:xVal><c:numRef><c:f>ダッシュボード!$%s$%d</c:f><c:numCache><c:formatCode>General</c:formatCode><c:ptCount val="1"/></c:numCache></c:numRef></c:xVal>' % (x_col, r)
                + '<c:yVal><c:numRef><c:f>ダッシュボード!$%s$%d</c:f><c:numCache><c:formatCode>General</c:formatCode><c:ptCount val="1"/></c:numCache></c:numRef></c:yVal>' % (y_col, r)
                + '<c:smooth val="0"/></c:ser>')
        c = c[:ser0.start()] + ''.join(body) + c[ser0.end():]
        c = c.replace('<c:varyColors val="1"/>', '<c:varyColors val="0"/>', 1)
        from collections import Counter
        return c, Counter(pos.values())

    for n, ncol, xcol, ycol, names in ((1, 'AH', 'AF', 'AG', names1), (2, 'AM', 'AK', 'AL', names2)):
        p = 'xl/charts/chart%d.xml' % n
        c = parts[p].decode('utf-8')
        assert len(re.findall(r'<c:ser>', c)) == 1
        c, cnt = rebuild(c, ncol, xcol, ycol, None, names)
        assert len(re.findall(r'<c:ser>', c)) == 60
        parts[p] = c.encode('utf-8')
        print('   chart%d: 60 系列に組み替え、ラベル位置 %s' % (n, dict(cnt)))
    print('chart1/2: 1人1系列（系列名＝AH/AM の氏名セル）に組み替え、X/Y は行ごとの単一セル参照に')

    zout = zipfile.ZipFile(DST, 'w', zipfile.ZIP_DEFLATED)
    for nm in order:
        zi = zipfile.ZipInfo(nm, date_time=infos[nm].date_time)
        zi.compress_type = infos[nm].compress_type
        zi.external_attr = infos[nm].external_attr
        zout.writestr(zi, parts[nm])
    zout.close()


main()
