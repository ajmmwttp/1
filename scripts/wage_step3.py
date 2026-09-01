# -*- coding: utf-8 -*-
"""段3+4+6: テーブル/判定行色を U列まで拡張し、集計行と使い方の説明を足す。"""
import re, sys, zipfile
import openpyxl

SRC, DST, CALC = sys.argv[1], sys.argv[2], sys.argv[3]
SHEETS = {'xl/worksheets/sheet4.xml': ('ピッキング', 'C$26', 'xl/tables/table1.xml'),
          'xl/worksheets/sheet5.xml': ('梱包', 'C$27', 'xl/tables/table2.xml')}
WAGE = 1500


def insert_cells(rowxml, newcells):
    m = re.match(r'(<row [^>]*>)(.*)(</row>)$', rowxml, re.S)
    head, body, tail = m.groups()
    cells = re.findall(r'<c r="[A-Z]+\d+"[^>]*/>|<c r="[A-Z]+\d+"[^>]*>.*?</c>', body, re.S)
    cells.extend(newcells)
    return head + ''.join(sorted(cells, key=lambda c: openpyxl.utils.column_index_from_string(
        re.match(r'<c r="([A-Z]+)', c).group(1)))) + tail


def main():
    calc = openpyxl.load_workbook(CALC, data_only=True)

    zin = zipfile.ZipFile(SRC)
    parts = {n: zin.read(n) for n in zin.namelist()}
    infos = {i.filename: i for i in zin.infolist()}
    order = zin.namelist()
    zin.close()

    for path, (name, ck, tpath) in SHEETS.items():
        s = parts[path].decode('utf-8')
        ws = calc[name]
        tgt = calc['設定'][ck.replace('$', '')].value

        # 判定の行色を U列まで
        assert '<conditionalFormatting sqref="A7:T66">' in s
        s = s.replace('<conditionalFormatting sqref="A7:T66">',
                      '<conditionalFormatting sqref="A7:U66">', 1)

        # 集計行（行20の V/W は空き）
        val = (ws['W12'].value - ws['W10'].value * tgt) / 60 * WAGE
        cells = [
            '<c r="V20" s="47" t="inlineStr"><is><t>目標到達で減らせる人件費(円)</t></is></c>',
            '<c r="W20" s="26"><f>IFERROR((W12-W10*設定!$%s)/60*設定!$C$28,"")</f><v>%s</v></c>'
            % (ck, repr(val)),
        ]
        m = re.search(r'<row r="20"[^>]*>.*?</row>', s, re.S)
        assert '<c r="V20"' not in m.group(0) and '<c r="W20"' not in m.group(0)
        s = s[:m.start()] + insert_cells(m.group(0), cells) + s[m.end():]
        parts[path] = s.encode('utf-8')

        # テーブル拡張
        t = parts[tpath].decode('utf-8')
        assert t.count('A6:T66') == 2
        t = t.replace('A6:T66', 'A6:U66')
        t = t.replace('<tableColumns count="20">', '<tableColumns count="21">')
        t = t.replace('</tableColumns>',
                      '<tableColumn id="21" name="目標到達で&#10;減らせる人件費(円)"/></tableColumns>')
        parts[tpath] = t.encode('utf-8')
        print('%s: 行色 A7:U66 / テーブル A6:U66・21列 / 集計行 {:+,.0f}円'.format(val) % name)

    # 使い方 C34（inlineStr）に金額列の説明を追記
    u = parts['xl/worksheets/sheet11.xml'].decode('utf-8')
    old = 'マイナスは、目標ペースより先行している分です。'
    assert u.count(old) == 1
    u = u.replace(old, old + '隣の人件費(円)は、この差の時間を時給換算の単価'
                             '（設定シート、現在1,500円/時）で金額にしたものです。')
    parts['xl/worksheets/sheet11.xml'] = u.encode('utf-8')
    print('使い方 C34: 金額列の説明を追記')

    zout = zipfile.ZipFile(DST, 'w', zipfile.ZIP_DEFLATED)
    for nm in order:
        zi = zipfile.ZipInfo(nm, date_time=infos[nm].date_time)
        zi.compress_type = infos[nm].compress_type
        zi.external_attr = infos[nm].external_attr
        zout.writestr(zi, parts[nm])
    zout.close()


main()
