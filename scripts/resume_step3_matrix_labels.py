# -*- coding: utf-8 -*-
"""段3: ③マトリクス（chart10）の氏名ラベルの重なりをさらに減らす。

前回（matrix_step2）は「両方の人: 梱包効率 100% 以上なら点の上、60% 未満も上、
その間は下」「帯の人: 必ず点の上」の規則で置いたが、描画して確かめると

  1) 帯（片方工程のみ）の人は PK 効率順に 6 段の帯に並ぶため、隣り合う段の人は
     横位置も近い。段の間隔は 6pt 相当しかなく、「点の上」に置いたラベルが
     すぐ上の段の点・ラベルに必ず重なる（上村和恵／秋草文恵／高橋正美／廣瀬こずえ）。
     → 帯のラベルは「点の右」に置く。段の高さ（約 11pt）の中に 7pt のラベルが
       収まるので、上下の段とは重ならない。同じ段に来るのは 6 人おきなので横も空く。
  2) 「60% 未満は上」の例外は、帯のラベルが上に伸びていた時の回避策なので不要になる。
     → 両方の人は「100% を境に外向き（以上なら上・未満なら下）」だけに単純化。
  3) それでも点同士が近い 2 組（山本祥吉・後藤克行・高橋佐織／木村千鶴・中納沙織）は
     規則だけでは重なる。
     → 規則の位置で衝突する点だけ、上下反対→右→左 の順に空いている側へ逃がす
       （逃がした人は実行時に表示する）。衝突判定は幅 350pt×高さ 340pt の
       プロット領域・7pt フォントという控えめな見積りで行う。実際の Excel の
       プロット領域はこれより横に広いので、ここで空いていれば実物でも空く。

象限ラベルの枠（ピッキング改善・良好・要改善（両方）・梱包改善）と
「▲ もう一方の工程は実績なし」の枠も障害物として避ける（菊池敬は下だと
「要改善（両方）」の枠にかかるので上に逃がす）。
"""
import itertools, re, sys, zipfile
import openpyxl

SRC, DST, VALS = sys.argv[1], sys.argv[2], sys.argv[3]
CHART = 'xl/charts/chart10.xml'

XMIN, XMAX, YMIN, YMAX = 40.0, 160.0, 10.0, 190.0
PW, PH = 350.0, 340.0                  # プロット領域の控えめな見積り(pt)
PX, PY = PW / (XMAX - XMIN), PH / (YMAX - YMIN)
FONT = 7.0
CW, LH, GAP, MR = FONT, FONT * 1.2, 2.5, 3.5
# 図形の枠（データ座標の近似）: 象限ラベル 4 枚と帯の注記
#   セルにアンカーされた図形なので、プロット領域との位置関係は Excel と LibreOffice で
#   多少ずれる。描画で測った値（左枠 x40〜74・下枠 y48〜60 など）に余裕を足してある。
OBSTACLES = [(40, 160, 76, 177), (110, 160, 155, 177),
             (40, 46, 76, 62), (110, 46, 155, 62),
             (128, 23, 160, 43)]


def to_pt(x, y):
    return (x - XMIN) * PX, (y - YMIN) * PY


def lbl_box(x, y, pos, name):
    cx, cy = to_pt(x, y)
    w, h = len(name) * CW, LH
    if pos == 't':
        return (cx - w / 2, cy + MR + GAP, cx + w / 2, cy + MR + GAP + h)
    if pos == 'b':
        return (cx - w / 2, cy - MR - GAP - h, cx + w / 2, cy - MR - GAP)
    if pos == 'r':
        return (cx + MR + GAP, cy - h / 2, cx + MR + GAP + w, cy + h / 2)
    return (cx - MR - GAP - w, cy - h / 2, cx - MR - GAP, cy + h / 2)


def mk_box(x, y):
    cx, cy = to_pt(x, y)
    return (cx - MR, cy - MR, cx + MR, cy + MR)


def obs_box(o):
    x0, y0 = to_pt(o[0], o[1])
    x1, y1 = to_pt(o[2], o[3])
    return (x0, y0, x1, y1)


def hit(a, b):
    return not (a[2] <= b[0] or b[2] <= a[0] or a[3] <= b[1] or b[3] <= a[1])


def inside(b):
    return b[0] >= -1 and b[2] <= PW + 1 and b[1] >= -1 and b[3] <= PH + 1


def rule(x, y, band):
    if band:
        return 'r'
    return 't' if y >= 100 else 'b'


def candidates(x, y, band):
    r = rule(x, y, band)
    if band:
        return [r, 'l', 't']
    return [r, 'b' if r == 't' else 't', 'r', 'l']


def solve(points):
    """points: {row: (name, x, y, band)} → {row: pos}。規則を優先し、衝突する点だけ逃がす。"""
    obst = [obs_box(o) for o in OBSTACLES]
    markers = {r: mk_box(p[1], p[2]) for r, p in points.items()}
    placed, out, moved = {}, {}, []
    for r in sorted(points, key=lambda k: (-points[k][2], points[k][1])):
        nm, x, y, band = points[r]
        chosen = None
        for pos in candidates(x, y, band):
            b = lbl_box(x, y, pos, nm)
            if not inside(b):
                continue
            if any(hit(b, o) for o in obst):
                continue
            if any(hit(b, mb) for k, mb in markers.items() if k != r):
                continue
            if any(hit(b, pb) for pb in placed.values()):
                continue
            chosen = pos
            break
        if chosen is None:
            chosen = rule(x, y, band)
        if chosen != rule(x, y, band):
            moved.append((nm, rule(x, y, band), chosen))
        out[r] = chosen
        placed[r] = lbl_box(x, y, chosen, nm)
    return out, moved, placed, markers


def count_collisions(placed, markers):
    ll = sum(1 for a, b in itertools.combinations(placed, 2) if hit(placed[a], placed[b]))
    lm = sum(1 for a in placed for b in markers if a != b and hit(placed[a], markers[b]))
    return ll, lm


def main():
    ws = openpyxl.load_workbook(VALS, data_only=True)['印刷用データ']
    points = {}
    for r in range(3, 63):
        nm = ws.cell(r, 33).value                      # AG
        if not nm:
            continue
        if ws.cell(r, 37).value == '両方':            # AK
            points[('both', r)] = (nm, float(ws.cell(r, 34).value), float(ws.cell(r, 35).value), False)
        else:
            points[('band', r)] = (nm, float(ws.cell(r, 49).value), float(ws.cell(r, 50).value), True)
    pos, moved, placed, markers = solve(points)
    ll, lm = count_collisions(placed, markers)
    print('対象 %d 名（両方 %d・帯 %d）' % (len(points),
          sum(1 for k in points if k[0] == 'both'), sum(1 for k in points if k[0] == 'band')))
    print('規則どおりに置けなかった人: %d 名' % len(moved))
    for nm, a, b in moved:
        print('   %s: %s → %s' % (nm, a, b))
    print('見積り上の残り重なり: ラベル同士 %d / ラベル×点 %d' % (ll, lm))

    zin = zipfile.ZipFile(SRC)
    parts = {n: zin.read(n) for n in zin.namelist()}
    infos = {i.filename: i for i in zin.infolist()}
    order = zin.namelist()
    zin.close()

    s = parts[CHART].decode('utf-8')
    out, pos_end, n_chg, n_ser = [], 0, 0, 0
    for m in re.finditer(r'<c:ser>(?:(?!</c:ser>).)*?</c:ser>', s, re.S):
        ser = m.group(0)
        out.append(s[pos_end:m.start()])
        pos_end = m.end()
        fs = re.findall(r'<c:f>([^<]*)</c:f>', ser)
        if len(fs) < 3:                                  # 基準線 2 本
            out.append(ser)
            continue
        n_ser += 1
        band = '$AW$' in fs[1]
        row = int(re.search(r'\$(\d+)$', fs[0]).group(1))
        want = pos.get(('band' if band else 'both', row))
        if want is None:                                 # データの無い行は規則の既定値
            want = 'r' if band else 't'
        mp = re.search(r'<c:dLblPos val="([a-z]+)"/>', ser)
        assert mp, '系列 %s に dLblPos がない' % fs[0]
        if mp.group(1) != want:
            ser = ser[:mp.start()] + '<c:dLblPos val="%s"/>' % want + ser[mp.end():]
            n_chg += 1
        out.append(ser)
    out.append(s[pos_end:])
    parts[CHART] = ''.join(out).encode('utf-8')
    print('chart10: 点の系列 %d 本のうち %d 本のラベル位置を変更' % (n_ser, n_chg))

    zout = zipfile.ZipFile(DST, 'w', zipfile.ZIP_DEFLATED)
    for nm in order:
        zi = zipfile.ZipInfo(nm, date_time=infos[nm].date_time)
        zi.compress_type = infos[nm].compress_type
        zi.external_attr = infos[nm].external_attr
        zout.writestr(zi, parts[nm])
    zout.close()


main()
