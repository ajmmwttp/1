# -*- coding: utf-8 -*-
"""案②（改善優先の上位N名に絞る）の N を決める実験。"""
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl

plt.rcParams["font.family"] = "IPAPGothic"
BLUE, BROWN, RED = "#2E75B6", "#C55A11", "#C00000"
W, H = 5.08, 3.5

def load(path):
    ws = openpyxl.load_workbook(path, data_only=True)["印刷用データ"]
    good = [(ws[f"AP{r}"].value, ws[f"AQ{r}"].value, ws[f"AR{r}"].value)
            for r in range(3, 63) if ws[f"AP{r}"].value]
    return list(reversed(good))          # 改善優先＝悪い順

def draw(ax, data, title):
    xs = range(len(data)); w = 0.4
    ax.bar([i-w/2 for i in xs], [d[1] or 0 for d in data], w, color=BLUE, label="ピッキング効率")
    ax.bar([i+w/2 for i in xs], [d[2] or 0 for d in data], w, color=BROWN, label="梱包効率")
    ax.axhline(100, color=RED, ls="--", lw=0.9, label="目標100%")
    for i, d in enumerate(data):
        for off, v in ((-w/2, d[1]), (w/2, d[2])):
            if v: ax.text(i+off, v+2, f"{v:.0f}", ha="center", fontsize=4.5, rotation=90)
    ax.set_xticks(list(xs)); ax.set_xticklabels([d[0] for d in data], rotation=60, ha="right", fontsize=5.5)
    ax.set_ylabel("効率（%）", fontsize=6); ax.tick_params(labelsize=5)
    ax.grid(axis="y", color="0.88", lw=0.5); ax.set_axisbelow(True)
    ax.legend(fontsize=5, ncol=3, loc="upper right")
    ax.set_title(title, fontsize=8.5, pad=4)

worst50 = load("chk_50.xlsx")
worst29 = load("chk_real.xlsx")

fig, axes = plt.subplots(2, 2, figsize=(W*2.15, H*2.15))
draw(axes[0][0], worst50[:15], "50名中 上位15名（1本 12.2pt）")
draw(axes[0][1], worst50[:20], "50名中 上位20名（1本 9.2pt）")
draw(axes[1][0], worst50[:25], "50名中 上位25名（1本 7.3pt）")
draw(axes[1][1], worst29[:20], "参考: 現在の実データ29名 → 上位20名")
fig.suptitle("案② 改善優先の上位N名に絞る — N の比較（実領域と同じ縦横比）", fontsize=11)
fig.tight_layout(rect=[0, 0, 1, 0.95])
fig.savefig("exp_topn.png", dpi=115)
print("wrote exp_topn.png")
