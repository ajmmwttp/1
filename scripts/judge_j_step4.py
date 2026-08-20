# -*- coding: utf-8 -*-
"""段4: 凡例（S6:S9）と集計（T6:T9）を4区分に更新する。"""
import re
import zipfile

import openpyxl
from collections import Counter

SRC, DST = "j3.xlsx", "j4.xlsx"
LABELS = [("◎ 達成", "◎達成"), ("○ あと一歩", "○あと一歩"),
          ("△ 要改善", "△要改善"), ("✕ 要支援", "✕要支援")]

wv = openpyxl.load_workbook(SRC, data_only=True)
COUNTS = {}
for sn, nm in [("sheet4", "ピッキング"), ("sheet5", "梱包")]:
    c = Counter(wv[nm].cell(r, 15).value for r in range(7, 67) if wv[nm].cell(r, 15).value)
    COUNTS[sn] = c
    print(nm, dict(c))

def patch(x, sn):
    for i, (legend, key) in enumerate(LABELS):
        r = 6 + i
        # 凡例 S列
        if r <= 8:
            m = re.search(r'<c r="S%d" s="47" t="s"><v>\d+</v></c>' % r, x)
        else:
            m = re.search(r'<c r="S9" s="47"/>', x)
        assert m, f"S{r} が想定と違う"
        x = x.replace(m.group(0),
                      f'<c r="S{r}" s="47" t="inlineStr"><is><t>{legend}</t></is></c>', 1)
        # 集計 T列
        n = COUNTS[sn].get(key, 0)
        new_t = f'<c r="T{r}" s="26"><f>COUNTIF(O7:O66,"{key}")</f><v>{n}</v></c>'
        if r <= 8:
            m = re.search(r'<c r="T%d" s="26"><f>COUNTIF\(O7:O66,"[^"]+"\)</f><v>\d+</v></c>' % r, x)
        else:
            m = re.search(r'<c r="T9" s="21"/>', x)
        assert m, f"T{r} が想定と違う"
        x = x.replace(m.group(0), new_t, 1)
    return x

zi = zipfile.ZipFile(SRC); zo = zipfile.ZipFile(DST, "w", zipfile.ZIP_DEFLATED)
for it in zi.infolist():
    b = zi.read(it.filename)
    if it.filename in ("xl/worksheets/sheet4.xml", "xl/worksheets/sheet5.xml"):
        b = patch(b.decode("utf-8"), it.filename.split("/")[-1][:-4]).encode("utf-8")
    zo.writestr(it, b)
zo.close(); zi.close()
print("wrote", DST)
