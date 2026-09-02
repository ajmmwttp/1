# -*- coding: utf-8 -*-
"""段3: 梱包側の取得を「行位置」から「入力行キー結合」に書き換える。

対象（各60行）:
  印刷用データ C(梱包K) F(梱包I) G(梱包D)   … 行3+i ↔ ピッキング行7+i
  散布図       Y(梱包L) Z(梱包E/C) AA(梱包G) … 行8+i ↔ ピッキング行7+i
  配置マップ   D(梱包L)                       … 行8+i ↔ ピッキング行7+i
共通パターン: X = INDEX(梱包!$列$7:$列$66, MATCH(ピッキング!$V$p, 梱包!$V$7:$V$66, 0))
             式 = IFERROR(IF(X="","",X), "")
"""
import re, sys, zipfile
import openpyxl

SRC, DST = sys.argv[1], sys.argv[2]


def ix(col, p):
    return ('INDEX(梱包!$%s$7:$%s$66,MATCH(ピッキング!$V$%d,梱包!$V$7:$V$66,0))'
            % (col, col, p))


def simple(col, p):
    x = ix(col, p)
    return 'IFERROR(IF(%s="","",%s),"")' % (x, x)


def ratio(p):
    e, c = ix('E', p), ix('C', p)
    return 'IFERROR(IF(OR(%s="",%s=""),"",%s/%s),"")' % (e, c, e, c)


# (worksheetパス, 対象列, 先頭行, 旧式に含まれるはずの語, 新式生成関数)
TARGETS = [
    ('xl/worksheets/sheet12.xml', 'C', 3, '梱包!$K$', lambda p: simple('K', p)),
    ('xl/worksheets/sheet12.xml', 'F', 3, '梱包!$I$', lambda p: simple('I', p)),
    ('xl/worksheets/sheet12.xml', 'G', 3, '梱包!$D$', lambda p: simple('D', p)),
    ('xl/worksheets/sheet7.xml', 'Y', 8, '梱包!$L$', lambda p: simple('L', p)),
    ('xl/worksheets/sheet7.xml', 'Z', 8, '梱包!$E$', ratio),
    ('xl/worksheets/sheet7.xml', 'AA', 8, '梱包!$G$', lambda p: simple('G', p)),
    ('xl/worksheets/sheet8.xml', 'D', 8, '梱包!$L$', lambda p: simple('L', p)),
]
COLMAP = {'C': 11, 'F': 9, 'G': 4, 'Y': 12, 'AA': 7, 'D': 12}   # 取得元の梱包列


def main():
    zin = zipfile.ZipFile(SRC)
    parts = {n: zin.read(n) for n in zin.namelist()}
    infos = {i.filename: i for i in zin.infolist()}
    order = zin.namelist()
    zin.close()

    # 正しいキー結合の値をPythonで用意
    wb = openpyxl.load_workbook(SRC, data_only=True)
    wf = openpyxl.load_workbook(SRC)
    pk, kp = wf['ピッキング'], wf['梱包']
    key_pk = {r: pk.cell(r, 22).value for r in range(7, 67)}    # V列
    kp_row_by_key = {kp.cell(r, 22).value: r for r in range(7, 67)}
    kpv = wb['梱包']

    def joined(col_idx, p):
        row = kp_row_by_key[key_pk[p]]
        return kpv.cell(row, col_idx).value

    for path, col, row0, expect, mk in TARGETS:
        s = parts[path].decode('utf-8')
        n = 0
        for i in range(60):
            r, p = row0 + i, 7 + i
            m = re.search(r'<c r="%s%d"([^>]*)>((?:(?!</c>).)*)</c>' % (col, r), s, re.S)
            assert m, '%s %s%d がない' % (path, col, r)
            attrs, body = m.group(1), m.group(2)
            mf = re.search(r'<f[^>]*>((?:(?!</f>).)*)</f>', body, re.S)
            assert mf and expect in mf.group(1), \
                '%s %s%d の式が想定と違う: %s' % (path, col, r, (mf.group(1) if mf else body)[:80])
            if col == 'Z':
                e, c = joined(5, p), joined(3, p)
                val = (e / c) if isinstance(e, (int, float)) and isinstance(c, (int, float)) and c else None
            else:
                v = joined(COLMAP[col], p)
                val = v if isinstance(v, (int, float)) else None
            attrs = re.sub(r' t="\w+"', '', attrs)
            if val is None:
                cell = '<c r="%s%d"%s t="str"><f>%s</f><v/></c>' % (col, r, attrs, mk(p))
            else:
                cell = '<c r="%s%d"%s><f>%s</f><v>%s</v></c>' % (col, r, attrs, mk(p), repr(val))
            s = s[:m.start()] + cell + s[m.end():]
            n += 1
        parts[path] = s.encode('utf-8')
        print('%s %s列: %d行をキー結合に書き換え' % (path.split('/')[-1], col, n))

    zout = zipfile.ZipFile(DST, 'w', zipfile.ZIP_DEFLATED)
    for nm in order:
        zi = zipfile.ZipInfo(nm, date_time=infos[nm].date_time)
        zi.compress_type = infos[nm].compress_type
        zi.external_attr = infos[nm].external_attr
        zout.writestr(zi, parts[nm])
    zout.close()


main()
