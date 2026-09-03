# -*- coding: utf-8 -*-
"""段2: 全チャートの埋め込みキャッシュ（strCache/numCache）を真値で更新する。

再計算しない閲覧環境（PDF変換など）でもグラフが正しく描かれるようにする。
系列・軸・書式などチャートの構造には一切触れず、<c:pt> の値だけ入れ替える。
"""
import re, sys, zipfile
import openpyxl
from openpyxl.utils import column_index_from_string as ci

SRC, DST, TRUTH = sys.argv[1], sys.argv[2], sys.argv[3]
CELL = re.compile(r'^\$?([A-Z]{1,3})\$?(\d+)$')


def xesc(t):
    return t.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


def resolve(wb, ref):
    """'シート!$A$3:$A$62' -> セル値のリスト。解決できなければ None。"""
    if '!' not in ref:
        return None
    sheet, rng = ref.rsplit('!', 1)
    sheet = sheet.strip("'").replace("''", "'")
    if sheet not in wb.sheetnames:
        return None
    ws = wb[sheet]
    parts = rng.split(':')
    m1 = CELL.match(parts[0])
    if not m1:
        return None
    if len(parts) == 1:
        return [ws.cell(int(m1.group(2)), ci(m1.group(1))).value]
    m2 = CELL.match(parts[1])
    if not m2:
        return None
    c1, r1 = ci(m1.group(1)), int(m1.group(2))
    c2, r2 = ci(m2.group(1)), int(m2.group(2))
    out = []
    for r in range(min(r1, r2), max(r1, r2) + 1):
        for c in range(min(c1, c2), max(c1, c2) + 1):
            out.append(ws.cell(r, c).value)
    return out


def build_cache(kind, vals, fmt):
    """strCache / numCache のXMLを組み立てる。"""
    pts = []
    for i, v in enumerate(vals):
        if v is None or v == '':
            continue
        if kind == 'num':
            if not isinstance(v, (int, float)) or isinstance(v, bool):
                continue
            pts.append('<c:pt idx="%d"><c:v>%s</c:v></c:pt>' % (i, repr(float(v))))
        else:
            pts.append('<c:pt idx="%d"><c:v>%s</c:v></c:pt>' % (i, xesc(str(v))))
    head = '<c:%sCache>' % kind
    if kind == 'num':
        head += fmt or '<c:formatCode>General</c:formatCode>'
    head += '<c:ptCount val="%d"/>' % len(vals)
    return head + ''.join(pts) + '</c:%sCache>' % kind


def main():
    truth = openpyxl.load_workbook(TRUTH, data_only=True)
    zin = zipfile.ZipFile(SRC)
    parts = {n: zin.read(n) for n in zin.namelist()}
    infos = {i.filename: i for i in zin.infolist()}
    order = zin.namelist()
    zin.close()

    charts = sorted([n for n in parts if re.match(r'xl/charts/chart\d+\.xml$', n)],
                    key=lambda x: int(re.search(r'(\d+)', x).group(1)))
    grand = 0
    for cp in charts:
        s = parts[cp].decode('utf-8')
        n = skipped = 0
        for kind in ('str', 'num'):
            pat = re.compile(
                r'(<c:f>([^<]*)</c:f>)(<c:%sCache>)((?:(?!</c:%sCache>).)*)(</c:%sCache>)'
                % (kind, kind, kind), re.S)

            def repl(m):
                nonlocal n, skipped
                vals = resolve(truth, m.group(2))
                if vals is None:
                    skipped += 1
                    return m.group(0)
                fmt = re.search(r'<c:formatCode>[^<]*</c:formatCode>', m.group(4))
                n += 1
                return m.group(1) + build_cache(kind, vals, fmt.group(0) if fmt else None)
            s = pat.sub(repl, s)
        parts[cp] = s.encode('utf-8')
        grand += n
        print('  %s: キャッシュ %d 個を更新%s'
              % (cp.split('/')[-1], n, '（未解決 %d）' % skipped if skipped else ''))
    print('合計 %d キャッシュ' % grand)

    zout = zipfile.ZipFile(DST, 'w', zipfile.ZIP_DEFLATED)
    for nm in order:
        zi = zipfile.ZipInfo(nm, date_time=infos[nm].date_time)
        zi.compress_type = infos[nm].compress_type
        zi.external_attr = infos[nm].external_attr
        zout.writestr(zi, parts[nm])
    zout.close()


main()
