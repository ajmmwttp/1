# -*- coding: utf-8 -*-
"""③改善優先度マトリクスに「片方の工程しかない人」を載せる設計の比較サンプル。

実データ（印刷用データ AG/AH/AI/AK）を読み、Excel に実装する前に
matplotlib で見え方を確認するための実験用スクリプト。
"""
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
import openpyxl

plt.rcParams["font.family"] = "IPAPGothic"
plt.rcParams["axes.unicode_minus"] = False

BLUE, GRAY, ORANGE, RED, GREEN = "#2E75B6", "#808080", "#C55A11", "#C00000", "#1BAF7A"

ws = openpyxl.load_workbook("cur.xlsx", data_only=True)["印刷用データ"]
both, only_pk, only_pack = [], [], []
for r in range(3, 63):
    nm = ws[f"AG{r}"].value
    if not nm:
        continue
    pk, pack, proc = ws[f"AH{r}"].value, ws[f"AI{r}"].value, ws[f"AK{r}"].value
    if proc == "両方":
        both.append((nm, pk, pack))
    elif pk is not None:
        only_pk.append((nm, pk))
    else:
        only_pack.append((nm, pack))
print(f"両工程 {len(both)} / PKのみ {len(only_pk)} / 梱包のみ {len(only_pack)}")


def quadrants(ax, lo, hi):
    ax.axvline(100, color="0.35", ls="--", lw=0.9)
    ax.axhline(100, color="0.35", ls="--", lw=0.9)
    for (x, y, t, c) in [(0.03, 0.95, "ピッキング改善", ORANGE), (0.72, 0.95, "良好", GREEN),
                         (0.03, 0.05, "要改善（両方）", RED), (0.72, 0.05, "梱包改善", ORANGE)]:
        ax.text(x, y, t, transform=ax.transAxes, fontsize=8, color=c, ha="left", va="center",
                bbox=dict(boxstyle="square,pad=0.35", fc="white", ec=c, lw=1))


def scatter_names(ax, pts, color, marker="o"):
    for nm, x, y in pts:
        ax.plot(x, y, marker, color=color, ms=6, zorder=3)
        ax.annotate(nm, (x, y), textcoords="offset points", xytext=(0, 6),
                    ha="center", fontsize=5.5, zorder=4)


fig, axes = plt.subplots(1, 3, figsize=(19.5, 6.6))

# ---- 案A: 軸を下へ広げ、専用の帯に置く --------------------------------
ax = axes[0]
ax.set_xlim(40, 190); ax.set_ylim(20, 190)
ax.add_patch(Rectangle((40, 20), 150, 20, fc="#F2F2F2", ec="0.7", lw=0.8, zorder=0))
ax.text(42, 30, "梱包の実績なし（ピッキングのみ）", fontsize=7, color="0.35", va="center", zorder=2)
quadrants(ax, 40, 190)
scatter_names(ax, both, BLUE)
scatter_names(ax, [(nm, pk, 30) for nm, pk in only_pk], GRAY, "^")
ax.set_title("案A  軸を下へ広げ、専用の帯に分離", fontsize=11)

# ---- 案B: 現行の軸のまま、最下段に置く --------------------------------
ax = axes[1]
ax.set_xlim(40, 190); ax.set_ylim(40, 190)
quadrants(ax, 40, 190)
scatter_names(ax, both, BLUE)
scatter_names(ax, [(nm, pk, 46) for nm, pk in only_pk], GRAY, "^")
ax.set_title("案B  軸はそのまま、最下段に灰色で置く", fontsize=11)

# ---- 案C: 梱包を「総合達成率」で代用（= y も実測値を持たせる）--------
ax = axes[2]
ax.set_xlim(40, 190); ax.set_ylim(40, 190)
quadrants(ax, 40, 190)
scatter_names(ax, both, BLUE)
scatter_names(ax, [(nm, pk, pk) for nm, pk in only_pk], GRAY, "^")
ax.set_title("案C  梱包の代わりに本人のPK効率を置く（対角線上）", fontsize=11)

for ax in axes:
    ax.set_xlabel("ピッキング効率（%）", fontsize=9)
    ax.set_ylabel("梱包効率（%）", fontsize=9)
    ax.grid(color="0.85", lw=0.6)
    ax.set_axisbelow(True)
    ax.tick_params(labelsize=8)

fig.suptitle("③ 改善優先度マトリクス — 片方の工程しかない10名（灰色▲）の置き方 3案", fontsize=13)
fig.tight_layout(rect=[0, 0, 1, 0.95])
fig.savefig("sample_scatter.png", dpi=110)
print("wrote sample_scatter.png")
